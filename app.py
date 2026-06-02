import os
import re
import sys
import datetime
import io
from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
import openpyxl

# 파이프라인 코어 모듈 임포트
import roadmap_agent_llm_new as roadmap_agent_llm
from rag_engine import initialize_rag, get_rag_engine
from concurrent.futures import ThreadPoolExecutor

sys.stdout.reconfigure(encoding='utf-8', line_buffering=True)

app = Flask(__name__)
CORS(app)

# 업로드 폴더 설정 - 상대 경로 지정
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# ====================================================================
# 서버 시작 시 ChromaDB RAG 벡터 인덱스 초기화
# 자사 보안 솔루션 리스트.xlsx → Markdown 세그먼트 → ChromaDB 벡터 임베딩
# ====================================================================
print("=" * 60, flush=True)
print("[시스템] ChromaDB RAG 벡터 인덱스 초기화 시작...", flush=True)
try:
    initialize_rag(force_rebuild=False)
    print("[시스템] ✅ RAG 벡터 인덱스 초기화 완료!", flush=True)
except Exception as e:
    print(f"[시스템] ⚠️ RAG 초기화 실패 (수동 재구축 필요): {e}", flush=True)
print("=" * 60, flush=True)

def get_company_name_from_file(original_filename, filepath):
    """체크리스트 파일명 또는 내부 셀 데이터에서 기업명 추출"""
    # 1. 원본 파일명에서 '년도 기업명_내부보안...' 패턴 추출
    # {년도} {기업명}_내부보안점검_상세체크리스트_{버전}_{작성일시}
    # 예: '2026년 KG제로인_내부보안점검...'
    match = re.search(r'년\s*([A-Za-z0-9가-힣\s]+?)(?:_|\s|내부보안)', original_filename)
    if match:
        name = match.group(1).strip()
        if name:
            return name
    
    # 2. 첫 번째 시트(표지)에서 KG ICT 로고 이외에 KG xxx 으로 되어 있는 고객사명 검색
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.worksheets[0]
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    # KG로 시작하는 한국어/영어/숫자 단어 탐색 (예: KG제로인, KG케미칼 등)
                    found_names = re.findall(r'KG[가-힣A-Za-z0-9]+', cell)
                    for val in found_names:
                        val_clean = val.strip()
                        if val_clean != "KGICT" and "ICT" not in val_clean:
                            return val_clean
                            
        # 백업: 시트 이름 중 표지 관련 시트의 상단 영역 직접 스캔
        for name in wb.sheetnames:
            if "01_KG그룹" in name or "01_" in name or "표지" in name:
                sh = wb[name]
                for r in range(1, 15):
                    for c in range(1, 10):
                        val = sh.cell(row=r, column=c).value
                        if val and isinstance(val, str):
                            found = re.findall(r'KG[가-힣A-Za-z0-9]+', val)
                            for v in found:
                                v_clean = v.strip()
                                if v_clean != "KGICT" and "ICT" not in v_clean:
                                    return v_clean
    except Exception as e:
        print(f"[시스템] 표지 시트 파싱 예외 발생: {e}", flush=True)
    
    return "고객사"

def make_safe_filename(original_filename):
    """특수문자를 정제하고 시간 정보를 접두어로 붙여 안전하고 식별 가능한 한글 파일명 생성"""
    import re
    # 한글, 영문, 숫자, 공백, 언더바, 대시, 마침표 이외의 문자 제거
    clean_name = re.sub(r'[^\w\s\.\-_가-힣]', '', original_filename)
    clean_name = clean_name.strip()
    # 공백이 여러 개인 것을 언더바로 정돈
    clean_name = re.sub(r'\s+', '_', clean_name)
    # 타임스탬프 접두사 추가
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{timestamp}_{clean_name}"

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "파일이 전송되지 않았습니다."}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"status": "error", "message": "선택된 파일이 없습니다."}), 400
        
    if not file.filename.endswith('.xlsx'):
        return jsonify({"status": "error", "message": "엑셀 파일(.xlsx)만 업로드할 수 있습니다."}), 400
        
    # 식별 가능한 안전한 실명 타임스탬프 파일명 적용
    safe_filename = make_safe_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_filename)
    
    try:
        file.save(filepath)
    except Exception as save_err:
        print(f"파일 물리적 저장 실패: {save_err}", flush=True)
        return jsonify({"status": "error", "message": f"서버 파일 쓰기 권한 및 경로 에러: {str(save_err)}"}), 500
        
    try:
        # 빨간색 글씨 결함 셀 정보 및 기존 매핑 데이터 파싱
        items = roadmap_agent_llm.parse_red_cells_from_checklist(filepath)
        company = get_company_name_from_file(file.filename, filepath)
        print(f"파일 업로드 완료: {file.filename}, 고객사명: {company}, 감지된 결함 수: {len(items)}건", flush=True)
        
        return jsonify({
            "status": "success",
            "company": company,
            "filename": file.filename,
            "filepath": filepath,
            "items": items
        })
    except Exception as e:
        print(f"파일 분석 실패: {e}", flush=True)
        return jsonify({"status": "error", "message": f"파일 분석 실패: {str(e)}"}), 500

@app.route('/api/upload_env', methods=['POST'])
def upload_env_file():
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "파일이 전송되지 않았습니다."}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"status": "error", "message": "선택된 파일이 없습니다."}), 400
        
    if not file.filename.endswith('.xlsx'):
        return jsonify({"status": "error", "message": "엑셀 파일(.xlsx)만 업로드할 수 있습니다."}), 400
        
    safe_filename = make_safe_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_filename)
    
    try:
        file.save(filepath)
    except Exception as save_err:
        print(f"사전환경조사서 저장 실패: {save_err}", flush=True)
        return jsonify({"status": "error", "message": f"서버 파일 저장 에러: {str(save_err)}"}), 500
        
    # 기업명 파싱 시도
    company = "고객사"
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = None
        for name in wb.sheetnames:
            if "KG그룹" in name:
                sheet = wb[name]
                break
        if not sheet:
            sheet = wb.worksheets[2] if len(wb.worksheets) > 2 else wb.active
        
        val = sheet.cell(row=5, column=2).value
        if val:
            company_candidate = str(val).strip()
            if "가족사명" not in company_candidate and company_candidate != "":
                company = company_candidate
            else:
                company = get_company_name_from_file(file.filename, filepath)
        else:
            company = get_company_name_from_file(file.filename, filepath)
    except Exception as e:
        print(f"사전환경조사서 기업명 추출 실패: {e}", flush=True)
        company = get_company_name_from_file(file.filename, filepath)
        
    print(f"사전환경조사서 업로드 완료: {file.filename}, 추출된 고객사명: {company}", flush=True)
    return jsonify({
        "status": "success",
        "company": company,
        "filename": file.filename,
        "filepath": filepath
    })

@app.route('/api/upload_asset', methods=['POST'])
def upload_asset_file():
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "파일이 전송되지 않았습니다."}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"status": "error", "message": "선택된 파일이 없습니다."}), 400
        
    if not file.filename.endswith('.xlsx'):
        return jsonify({"status": "error", "message": "엑셀 파일(.xlsx)만 업로드할 수 있습니다."}), 400
        
    safe_filename = make_safe_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_filename)
    
    try:
        file.save(filepath)
    except Exception as save_err:
        print(f"자산목록 저장 실패: {save_err}", flush=True)
        return jsonify({"status": "error", "message": f"서버 파일 저장 에러: {str(save_err)}"}), 500
        
    print(f"자산목록 업로드 완료: {file.filename}", flush=True)
    return jsonify({
        "status": "success",
        "filename": file.filename,
        "filepath": filepath
    })

@app.route('/api/map', methods=['POST'])
def map_solutions():
    print("Received /api/map request!", flush=True)
    data = request.json
    items = data.get("items", [])
    model_name = data.get("model", "gemma2:2b")
    print(f"Items count to process: {len(items)}, selected model: {model_name}", flush=True)
    if not items:
        return jsonify({"status": "error", "message": "분석할 체크리스트 항목이 없습니다."}), 400
        
    def process_item(item):
        print(f"Processing row {item.get('row_idx')}: {item.get('항목명')}", flush=True)
        try:
            # 로컬 Ollama 호출 (sol_context=None으로 주면 아이템별 초경량 RAG 필터링 수행)
            map_res = roadmap_agent_llm.call_local_gemma(item, None, model_name=model_name)
            print(f"Ollama mapping result for row {item.get('row_idx')}: {map_res}", flush=True)
            
            # 연차 배분 스케줄링
            year = roadmap_agent_llm.calculate_roadmap_year(map_res)
            
            return {
                "row_idx": item["row_idx"],
                "항목명": item["항목명"],
                "세부점검내용": item["세부점검내용"],
                "운영현황_증적": item["운영현황_증적"],
                "개선방안": item["개선방안"],
                "보안영역": map_res.get("보안영역", "기타 보안"),
                "과제명": map_res.get("과제명", "보안 솔루션 구축"),
                "법적요구": map_res.get("법적요구", "N/A"),
                "시급성": map_res.get("시급성", 3),
                "위험도": map_res.get("위험도", 3),
                "예상예산": map_res.get("예상예산", "영업 문의가 필요한 영역 논의 필요."),
                "로드맵연도": year,
                "비고": map_res.get("비고", ""),
                "추천솔루션": map_res.get("추천솔루션", "N/A"),
                "제조사": map_res.get("제조사", "N/A")
            }
        except Exception as e:
            print(f"아이템 처리 중 에러 발생: {e}", flush=True)
            fallback_res = {
                "보안영역": "미정",
                "과제명": "솔루션 도입 검토 필요 (분석 실패)",
                "법적요구": "N/A",
                "시급성": 3,
                "위험도": 3,
                "예상예산": "영업 문의가 필요한 영역 논의 필요.",
                "비고": f"[추천 솔루션] 분석 실패\n[선정 이유] LLM 호출 도중 예외가 발생했습니다: {str(e)}",
                "추천솔루션": "N/A",
                "제조사": "N/A"
            }
            year = roadmap_agent_llm.calculate_roadmap_year(fallback_res)
            return {
                "row_idx": item["row_idx"],
                "항목명": item["항목명"],
                "세부점검내용": item["세부점검내용"],
                "운영현황_증적": item["운영현황_증적"],
                "개선방안": item["개선방안"],
                "보안영역": fallback_res["보안영역"],
                "과제명": fallback_res["과제명"],
                "법적요구": fallback_res["법적요구"],
                "시급성": fallback_res["시급성"],
                "위험도": fallback_res["위험도"],
                "예상예산": fallback_res["예상예산"],
                "로드맵연도": year,
                "비고": fallback_res["비고"],
                "추천솔루션": fallback_res["추천솔루션"],
                "제조사": fallback_res["제조사"]
            }

    # ThreadPoolExecutor를 사용한 병렬 처리 (최대 4개 스레드로 동시 요청 처리)
    with ThreadPoolExecutor(max_workers=4) as executor:
        mapped_results = list(executor.map(process_item, items))
    
    print("Finished processing all items.", flush=True)
        
    return jsonify({
        "status": "success",
        "results": mapped_results
    })

@app.route('/api/export', methods=['POST'])
def export_excel():
    print("Received /api/export request!", flush=True)
    
    # JSON 요청과 일반 Form Submit 방식을 모두 처리할 수 있게 범용화
    if request.is_json:
        data = request.json
    else:
        import json
        form_data_str = request.form.get("data", "{}")
        try:
            data = json.loads(form_data_str)
        except Exception as je:
            print(f"Form data JSON parsing error: {je}", flush=True)
            data = {}
            
    results = data.get("results", [])
    company = data.get("company", "고객사")
    env_filepath = data.get("env_filepath", None)
    asset_filepath = data.get("asset_filepath", None)
    
    if not results:
        print("내보내기 실패: 결과 결과 리스트가 비어있음", flush=True)
        return jsonify({"status": "error", "message": "추출할 데이터가 없습니다."}), 400
        
    # 날짜 포맷팅 및 파일명 지정
    # 포맷: {올해}년 KG그룹_{기업명} 정보보안감사_보안솔루션로드맵_{생성일}.xlsx
    year = datetime.datetime.now().strftime("%Y")
    today = datetime.datetime.now().strftime("%Y%m%d")
    filename = f"{year}년 KG그룹_{company} 정보보안감사_보안솔루션로드맵_{today}.xlsx"
    
    # 임시 파일 경로
    temp_output_path = os.path.join(app.config['UPLOAD_FOLDER'], f"temp_{today}_{company}.xlsx")
    
    try:
        # 서식 보존 엑셀 생성
        roadmap_agent_llm.generate_roadmap_excel(results, company, temp_output_path, env_filepath=env_filepath, asset_filepath=asset_filepath)
        print(f"임시 엑셀 생성 완료: {temp_output_path}", flush=True)
        
        # 바이너리 메모리 로딩 후 삭제 전송
        return_data = io.BytesIO()
        with open(temp_output_path, 'rb') as f:
            return_data.write(f.read())
        return_data.seek(0)
        
        # 임시 파일 물리적 삭제
        if os.path.exists(temp_output_path):
            os.remove(temp_output_path)
            
        print("엑셀 파일 전송 성공!", flush=True)
        response = send_file(
            return_data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
        # 브라우저 보안 다운로드 경고(Mixed Content, Untrusted Origin 등) 방지를 위한 헤더 강화
        response.headers['X-Content-Type-Options'] = 'nosniff'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    except Exception as e:
        print(f"엑셀 내보내기 에러 상세 로그:", flush=True)
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"엑셀 다운로드 실패: {str(e)}"}), 500

@app.route('/api/rag/rebuild', methods=['POST'])
def rebuild_rag():
    """ChromaDB RAG 벡터 인덱스 강제 재구축 API"""
    print("[API] RAG 인덱스 재구축 요청 수신", flush=True)
    try:
        engine = initialize_rag(force_rebuild=True)
        count = engine._get_collection().count()
        print(f"[API] ✅ RAG 인덱스 재구축 완료! 총 {count}건", flush=True)
        return jsonify({
            "status": "success",
            "message": f"RAG 벡터 인덱스 재구축 완료. 총 {count}건의 솔루션이 임베딩되었습니다."
        })
    except Exception as e:
        print(f"[API] RAG 재구축 실패: {e}", flush=True)
        return jsonify({"status": "error", "message": f"RAG 재구축 실패: {str(e)}"}), 500

@app.route('/api/rag/status', methods=['GET'])
def rag_status():
    """ChromaDB RAG 벡터 인덱스 상태 확인 API"""
    try:
        engine = get_rag_engine()
        collection = engine._get_collection()
        count = collection.count()
        return jsonify({
            "status": "success",
            "indexed_count": count,
            "initialized": engine._initialized
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/rag/clear', methods=['POST'])
def clear_rag():
    """ChromaDB RAG 벡터 DB 데이터 완전히 비우기 API"""
    print("[API] RAG 데이터 삭제 요청 수신", flush=True)
    try:
        engine = get_rag_engine()
        success = engine.clear_index()
        if success:
            return jsonify({
                "status": "success",
                "message": "ChromaDB RAG 벡터 데이터베이스가 성공적으로 삭제/비워졌습니다."
            })
        else:
            raise Exception("RAG 인덱스 비우기 과정 중 오류 발생")
    except Exception as e:
        print(f"[API] RAG 데이터 삭제 실패: {e}", flush=True)
        return jsonify({"status": "error", "message": f"RAG 삭제 실패: {str(e)}"}), 500

@app.route('/api/uploads/clear', methods=['POST'])
def clear_uploads():
    """uploads 폴더 내부의 모든 임시 파일 삭제 API"""
    print("[API] 업로드 임시 파일 일괄 삭제 요청 수신", flush=True)
    try:
        import glob
        files = glob.glob(os.path.join(app.config['UPLOAD_FOLDER'], "*"))
        removed_count = 0
        for f in files:
            if os.path.isfile(f):
                os.remove(f)
                removed_count += 1
        print(f"[API] 임시 파일 {removed_count}건 삭제 완료.", flush=True)
        return jsonify({
            "status": "success",
            "message": f"업로드된 임시 파일 {removed_count}건이 안전하게 영구 삭제되었습니다."
        })
    except Exception as e:
        print(f"[API] 임시 파일 삭제 실패: {e}", flush=True)
        return jsonify({"status": "error", "message": f"임시 파일 삭제 실패: {str(e)}"}), 500

@app.route('/api/uploads/list', methods=['GET'])
def list_uploads():
    """uploads 폴더 내의 파일 목록 반환 API"""
    try:
        import glob
        files = glob.glob(os.path.join(app.config['UPLOAD_FOLDER'], "*"))
        file_list = []
        for f in files:
            if os.path.isfile(f):
                stat = os.stat(f)
                file_list.append({
                    "filename": os.path.basename(f),
                    "size": stat.st_size,
                    "created_at": datetime.datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
        return jsonify({
            "status": "success",
            "files": file_list
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/rag/list', methods=['GET'])
def list_rag():
    """ChromaDB RAG 벡터 DB 전체 데이터 목록 반환 API"""
    try:
        engine = get_rag_engine()
        collection = engine._get_collection()
        data = collection.get(include=["metadatas", "documents"])
        
        results = []
        ids = data.get("ids", [])
        metadatas = data.get("metadatas", [])
        documents = data.get("documents", [])
        
        for idx in range(len(ids)):
            meta = metadatas[idx] if idx < len(metadatas) else {}
            doc = documents[idx] if idx < len(documents) else ""
            results.append({
                "id": ids[idx],
                "보안영역": meta.get("보안영역", "N/A"),
                "솔루션구분": meta.get("솔루션구분", "N/A"),
                "제조사명": meta.get("제조사명", "N/A"),
                "제품명": meta.get("제품명", "N/A"),
                "제품명기능설명": meta.get("제품명기능설명", "")[:200] + ("..." if len(meta.get("제품명기능설명", "")) > 200 else "")
            })
        return jsonify({
            "status": "success",
            "count": len(results),
            "data": results
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500



if __name__ == '__main__':
    print("Starting Flask server on https://localhost:5000 (with SSL adhoc context)", flush=True)
    # Vercel(HTTPS)과의 브라우저 Mixed Content 보안 차단 우회를 위해 adhoc SSL 컨텍스트 사용
    app.run(host='0.0.0.0', port=5000, ssl_context='adhoc', debug=False)
