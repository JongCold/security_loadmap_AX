import os
import re
import sys
import json
from copy import copy
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import google.generativeai as genai
from dotenv import load_dotenv

# 표준 출력 인코딩을 UTF-8로 강제 설정
sys.stdout.reconfigure(encoding='utf-8')

# .env 파일 로드
load_dotenv()

# 경로 정의
CHECKLIST_PATH = r"c:\Security loadmap_Auto\2026년 KG제로인_내부보안점검_상세체크리스트_테스트 파일.xlsx"
SOLUTION_LIST_PATH = r"c:\Security loadmap_Auto\자사 보안 솔루션 리스트 (1).xlsx"
ROADMAP_TEMPLATE_PATH = r"c:\Security loadmap_Auto\2026년 KG그룹_제로인 정보보안감사_보안솔루션로드맵_v1.2.xlsx"
OUTPUT_PATH = r"c:\Security loadmap_Auto\2026년 KG그룹_제로인 정보보안감사_보안솔루션로드맵_결과.xlsx"

# 룰 기반 매핑을 위한 키워드 사전 정의
RULE_MAPPING_RULES = [
    {
        "keywords": ["2FA", "OTP", "추가 인증", "다중 인증", "인증 및 권한", "MFA", "추가인증", "다중인증"],
        "sol_name": "OKTA",
        "vendor": "OKTA",
        "area": "사용자 보안",
        "budget": 20000000,
        "law": "개인정보 보호법 제29조(안전조치의무)",
        "desc": "MFA 도입을 통한 계정 도용 방지 및 인증 보안 강화"
    },
    {
        "keywords": ["방화벽", "FW", "IPS", "침입 방지", "외부 트래픽", "침입방지", "외부트래픽", "망 분리", "망분리"],
        "sol_name": "NGFW A(차세대 방화벽)",
        "vendor": "Fortinet",
        "area": "네트워크 보안",
        "budget": 45000000,
        "law": "개인정보 보호법 제29조(안전조치의무)",
        "desc": "차세대 방화벽 도입 및 강력한 외부 트래픽 통제 규칙 적용"
    },
    {
        "keywords": ["웹 방화벽", "웹방화벽", "WAF", "AIWAF", "웹 보안", "웹보안"],
        "sol_name": "AIWAF-500Y",
        "vendor": "MonitorLab",
        "area": "네트워크 보안",
        "budget": 30000000,
        "law": "개인정보 보호법 제29조(안전조치의무)",
        "desc": "웹 서비스 취약점 및 악성 공격 실시간 탐지/차단 전용 웹 방화벽 구축"
    },
    {
        "keywords": ["DB접근제어", "DBMS", "쿼리", "데이터베이스 접근", "데이터베이스접근", "DB 접근제어", "DB 접근"],
        "sol_name": "DB접근제어 (Chakra Max)",
        "vendor": "WareValley",
        "area": "시스템 보안",
        "budget": 35000000,
        "law": "개인정보 보호법 제29조(안전조치의무)",
        "desc": "핵심 DBMS 접근 권한 제어 및 실시간 감사 로그 관리 솔루션 도입"
    },
    {
        "keywords": ["접근제어", "서버 접근", "서버접근", "시스템 접근", "시스템접근", "접근 권한", "계정 관리", "계정관리"],
        "sol_name": "Hiware 시스템접근제어",
        "vendor": "Netand",
        "area": "시스템 보안",
        "budget": 40000000,
        "law": "개인정보 보호법 제29조(안전조치의무)",
        "desc": "서버 및 인프라 시스템 통합 접근제어 및 작업 이력 감사 환경 수립"
    },
    {
        "keywords": ["백신", "바이러스", "악성코드", "V3", "알약", "실시간 탐지"],
        "sol_name": "백신V3",
        "vendor": "Ahnlab",
        "area": "시스템 보안",
        "budget": 5000000,
        "law": "개인정보 보호법 제29조(안전조치의무)",
        "desc": "서버 및 엔드포인트 단말기의 신종 악성코드 실시간 차단 시스템 확보"
    },
    {
        "keywords": ["DLP", "매체제어", "오피스키퍼", "정보유출", "정보 유출", "USB 통제", "출력 통제"],
        "sol_name": "오피스키퍼",
        "vendor": "지란지교소프트",
        "area": "PC보안",
        "budget": 12000000,
        "law": "개인정보 보호법 제29조(안전조치의무)",
        "desc": "중요 영업 비밀 및 개인정보 유출 원천 차단을 위한 통합 PC 보안 솔루션 도입"
    },
    {
        "keywords": ["백업", "재해복구", "소실", "Veeam", "이중화", "데이터 복구"],
        "sol_name": "Veeam Backup & Replication",
        "vendor": "Veeam",
        "area": "데이터 보안",
        "budget": 25000000,
        "law": "개인정보 보호법 제29조(안전조치의무)",
        "desc": "주기적 백업 자동화 및 불의의 시스템 재해 상황 대비 이중화 백업 시스템 구축"
    },
    {
        "keywords": ["개인정보 검출", "Safer", "개인정보 필터", "개인정보검출", "Safer", "필터링"],
        "sol_name": "U-Privacy Safer",
        "vendor": "Somansa",
        "area": "데이터 보안",
        "budget": 18000000,
        "law": "개인정보 보호법 제29조(안전조치의무)",
        "desc": "네트워크 및 엔드포인트 상의 미승인 개인정보 파일 실시간 탐지/격리/암호화"
    },
    {
        "keywords": ["망연계", "망 연계", "i-oneNetDD", "망간 자료전송", "망간전송"],
        "sol_name": "i-oneNetDD",
        "vendor": "한싹",
        "area": "네트워크 보안",
        "budget": 30000000,
        "law": "개인정보 보호법 제29조(안전조치의무)",
        "desc": "인터넷망과 업무망 간의 안전한 데이터 단방향 자료전송 환경 구축"
    },
    {
        "keywords": ["컨설팅", "ISMS", "인증", "규정", "지침", "정책", "취약점 진단", "취약점진단"],
        "sol_name": "ISMS-P 컨설팅",
        "vendor": "자사 전문인력",
        "area": "보안인증 및 관리",
        "budget": 15000000,
        "law": "개인정보 보호법 제29조 / 정보통신망법 제47조",
        "desc": "정보보호 관리체계(ISMS-P) 수립 및 법규 준수성 자체 진단 컨설팅 지원"
    }
]

def load_security_solutions():
    """자사 보안 솔루션 리스트 엑셀 파일을 읽어서 DataFrame으로 변환"""
    print("[1/6] 자사 보안 솔루션 리스트 DB 로딩 중...")
    if not os.path.exists(SOLUTION_LIST_PATH):
        print(f"오류: 자사 보안 솔루션 리스트 파일이 존재하지 않습니다: {SOLUTION_LIST_PATH}")
        return None
    try:
        df = pd.read_excel(SOLUTION_LIST_PATH, sheet_name="Sheet1")
        # 컬럼 이름 정규화
        df.columns = [c.strip() for c in df.columns]
        print(f"성공: {len(df)}개의 자사 보안 솔루션 정보를 성공적으로 불러왔습니다.")
        return df
    except Exception as e:
        print(f"오류: 자사 보안 솔루션 리스트 로딩 실패: {e}")
        return None

def build_markdown_context(sol_df):
    """자사 솔루션 리스트를 LLM 주입용 Markdown 컨텍스트 문자열로 변환"""
    if sol_df is None:
        return ""
    md_lines = []
    md_lines.append("# [자사 보유 보안 솔루션 명세 라이브러리]")
    for idx, row in sol_df.iterrows():
        area = row.get("보안영역", "기타 보안")
        sol_type = row.get("솔루션 구분", "N/A")
        vendor = row.get("제조사명", "N/A")
        prod_name = row.get("제품명", "N/A")
        desc = row.get("제품명 기능 설명", "N/A")
        
        md_lines.append(f"## [{area}] {sol_type} (제품명: {prod_name} / 제조사: {vendor})")
        md_lines.append(f"- **기능 및 특징 상세**: {desc}")
        md_lines.append("")
    return "\n".join(md_lines)

def load_checklist():
    """입력 체크리스트 엑셀 파싱하여 JSON 친화적인 딕셔너리 리스트로 변환"""
    print("[2/6] 내부보안감사체크리스트 파싱 중...")
    if not os.path.exists(CHECKLIST_PATH):
        print(f"오류: 체크리스트 파일이 존재하지 않습니다: {CHECKLIST_PATH}")
        return []
    
    try:
        wb = openpyxl.load_workbook(CHECKLIST_PATH, data_only=True)
        # 4번 시트 (이름에 '내부보안감사체크리스트'가 포함되어 있거나 4번째 시트)
        sheet_name = None
        for name in wb.sheetnames:
            if "내부보안감사체크리스트" in name:
                sheet_name = name
                break
        if not sheet_name:
            sheet_name = wb.sheetnames[3] # 4번째 시트 인덱스
            
        sheet = wb[sheet_name]
        data = []
        
        # Row 2 (0-indexed 2)부터 시작하여 데이터를 읽어 들임 (Row 0, 1은 헤더)
        for r_idx in range(3, sheet.max_row + 1):
            item_name = sheet.cell(row=r_idx, column=3).value # 항목명 (Col 2, 0-indexed -> Col 3, 1-indexed)
            detail = sheet.cell(row=r_idx, column=4).value # 세부점검내용 (Col 3 -> Col 4)
            eval_val = sheet.cell(row=r_idx, column=5).value # 평가 (Col 4 -> Col 5)
            status = sheet.cell(row=r_idx, column=6).value # 운영현황 및 증적 (Col 5 -> Col 6)
            improvement = sheet.cell(row=r_idx, column=7).value # 개선방안 (Col 6 -> Col 7)
            
            # 개선방안이 비어있지 않거나 평가가 X인 행들을 매핑 대상으로 수집
            if item_name and (eval_val == "X" or (improvement and len(str(improvement).strip()) > 5)):
                data.append({
                    "row_idx": r_idx,
                    "항목명": str(item_name).strip(),
                    "세부점검내용": str(detail).strip() if detail else "",
                    "평가": str(eval_val).strip() if eval_val else "",
                    "운영현황_증적": str(status).strip() if status else "",
                    "개선방안": str(improvement).strip() if improvement else ""
                })
        
        print(f"성공: 총 {len(data)}건의 개선 대상 취약 및 점검 항목이 검출되었습니다.")
        return data
    except Exception as e:
        print(f"오류: 체크리스트 파싱 중 에러 발생: {e}")
        return []

def rule_based_mapping(item):
    """API KEY가 없거나 호출 에러 시 Fallback으로 작동하는 룰 기반 매핑 엔진"""
    text = (item["개선방안"] + " " + item["세부점검내용"] + " " + item["항목명"]).lower()
    
    # 기본값 설정
    best_rule = {
        "sol_name": "자사 보안 솔루션 검토",
        "vendor": "기타 벤더",
        "area": "기타 보안",
        "budget": 25000000,
        "law": "N/A",
        "desc": "개선방안에 따른 보안 솔루션 도입 권고"
    }
    
    # 키워드 매칭 수행
    max_matches = 0
    for rule in RULE_MAPPING_RULES:
        matches = sum(1 for kw in rule["keywords"] if kw.lower() in text)
        if matches > max_matches:
            max_matches = matches
            best_rule = rule
            
    # 시급성, 위험도 룰 기반 산정
    urgency = 3
    risk = 3
    
    # 텍스트 강조 표현이 있으면 가중치
    if any(x in text for x in ["필수", "시급", "법적", "규정", "준수", "위반", "과태료"]):
        urgency = 5
        risk = 4
    if any(x in text for x in ["미흡", "불량", "취약", "노출"]):
        risk = 5
        
    law_req = best_rule["law"]
    if law_req != "N/A":
        urgency = 5 # 법적 의무 사항은 시급성 5로 상향 고정
        
    return {
        "보안영역": best_rule["area"],
        "과제명": f"{best_rule['sol_name']} 도입 및 보안 통제 강화",
        "법적요구": law_req,
        "시급성": urgency,
        "위험도": risk,
        "예상예산": f"₩{best_rule['budget']:,}",
        "비고": f"[추천 솔루션] {best_rule['vendor']} - {best_rule['sol_name']}\n[선정 이유] {best_rule['desc']}"
    }

def llm_based_mapping(item, sol_context):
    """Gemini API를 이용한 생성적 RAG 매핑 엔진"""
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key or "your_gemini_api_key" in api_key:
        raise ValueError("유효한 GEMINI_API_KEY 환경변수가 존재하지 않습니다.")
        
    genai.configure(api_key=api_key)
    
    # 구조화된 출력(JSON)을 위한 프롬프트 작성
    prompt = f"""
역할: 정보보안 CISO 컨설턴트 및 솔루션 아키텍트.
주어진 [고객사 개선요구사항]을 기반으로, [자사 보유 보안 솔루션 명세 라이브러리]에서 가장 적절한 솔루션을 매핑하고 결과를 구조화된 JSON으로 반환하라.

[자사 보유 보안 솔루션 명세 라이브러리]
{sol_context}

[고객사 개선요구사항]
- 항목명: {item['항목명']}
- 세부점검내용: {item['세부점검내용']}
- 운영현황 및 증적: {item['운영현황_증적']}
- 개선방안: {item['개선방안']}

[출력 요구사항 및 JSON 스키마]
아래 명세 스펙의 키 이름을 가진 단일 JSON 객체 형태로만 대답하라. (마크다운 코드 블록이나 주석, 기타 설명 텍스트는 절대 포함하지 말고 순수 JSON 문자열만 출력해야 한다.)
{{
  "보안영역": "자사 솔루션 라이브러리의 '보안영역' 중 하나 기재 (예: 네트워크 보안, 시스템 보안, 데이터 보안, PC보안 등)",
  "과제명": "[제품명] 도입 및 보안 통제 강화 (예: NGFW A(차세대 방화벽) 도입 및 보안 통제 강화)",
  "법적요구": "개인정보보호법 제29조 등 관련 법적 요건 명시 (없을 시 'N/A')",
  "시급성": 1~5 사이의 정수값 (법적 요구사항이 존재하는 경우 무조건 5로 설정),
  "위험도": 1~5 사이의 정수값 (위협 강도 기준),
  "예상예산": "₩30,000,000" 형태의 문자열 (솔루션 규모 및 단가 고려)",
  "비고": "[추천 솔루션] 제조사명 - 제품명 기재\\n[선정 이유] 해당 솔루션이 개선방안을 어떻게 구체적으로 해결할 수 있는지 서술"
}}
"""
    try:
        model = genai.GenerativeModel("gemini-1.5-flash")
        response = model.generate_content(
            prompt,
            generation_config={"response_mime_type": "application/json"}
        )
        result = json.loads(response.text.strip())
        return result
    except Exception as e:
        print(f"경고: LLM API 호출 실패 ({e}), 룰 기반 엔진으로 자동 전환합니다.")
        return rule_based_mapping(item)

def calculate_roadmap_year(mapping_res):
    """시급성, 위험도, 법적요구 기반 로드맵 연도(2026 ~ 2028) 산정 스케줄러"""
    urgency = int(mapping_res.get("시급성", 3))
    risk = int(mapping_res.get("위험도", 3))
    law = mapping_res.get("법적요구", "N/A")
    
    score = urgency * risk
    
    # 1. 법적 요구가 존재하면 1년차(2026년) 강제 최우선 배치
    if law and law != "N/A":
        return "2026년"
        
    # 2. 스코어 매트릭스에 따른 분배
    if score >= 16:
        return "2026년"
    elif score >= 8:
        return "2027년"
    else:
        return "2028년"

def copy_cell_style(src_cell, dest_cell):
    """원본 셀의 스타일 서식 속성을 대상 셀로 완전 복사"""
    if src_cell.font:
        dest_cell.font = copy(src_cell.font)
    if src_cell.border:
        dest_cell.border = copy(src_cell.border)
    if src_cell.fill:
        dest_cell.fill = copy(src_cell.fill)
    if src_cell.alignment:
        dest_cell.alignment = copy(src_cell.alignment)
    if src_cell.number_format:
        dest_cell.number_format = copy(src_cell.number_format)

def write_to_roadmap_template(results):
    """최종 생성된 로드맵 JSON 데이터를 원본 엑셀 템플릿의 스타일을 보존하며 작성"""
    print("[5/6] 로드맵 결과 엑셀 작성 중...")
    if not os.path.exists(ROADMAP_TEMPLATE_PATH):
        print(f"오류: 로드맵 템플릿 파일이 존재하지 않습니다: {ROADMAP_TEMPLATE_PATH}")
        return False
        
    try:
        wb = openpyxl.load_workbook(ROADMAP_TEMPLATE_PATH)
        sheet_name = None
        for name in wb.sheetnames:
            if "01_KG그룹" in name or "01_" in name:
                sheet_name = name
                break
        if not sheet_name:
            sheet_name = wb.sheetnames[1] # 01_KG그룹(KG제로인) 시트 위치
            
        sheet = wb[sheet_name]
        
        # 엑셀의 데이터 작성이 시작되는 기준 행 (25행: 0-indexed 24행)
        # 로드맵 데이터 양식 상의 행 번호는 26행(1-indexed)부터가 안전한 데이터 기입 구역입니다.
        start_row = 26
        
        # 기존 26행(템플릿 스타일 보존용 원본 셀들)
        style_source_row = start_row
        
        for idx, res in enumerate(results):
            current_row = start_row + idx
            
            # 스타일을 적용하기 위해 행을 복제하듯이 작성
            for col_idx in range(1, 17): # Col 1 ~ Col 16 (No.부터 비고까지)
                src_cell = sheet.cell(row=style_source_row, column=col_idx)
                dest_cell = sheet.cell(row=current_row, column=col_idx)
                copy_cell_style(src_cell, dest_cell)
            
            # 값 할당
            sheet.cell(row=current_row, column=1, value=idx + 1) # No.
            sheet.cell(row=current_row, column=2, value=res["항목명"]) # 항목명 (체크리스트)
            sheet.cell(row=current_row, column=3, value=res["세부점검내용"]) # 세부점검내용 (체크리스트)
            sheet.cell(row=current_row, column=5, value=res["운영현황_증적"]) # 운영현황 및 증적 (체크리스트)
            sheet.cell(row=current_row, column=7, value=res["개선방안"]) # 개선방안 (체크리스트)
            
            sheet.cell(row=current_row, column=9, value=res["보안영역"]) # 보안영역 (로드맵)
            sheet.cell(row=current_row, column=10, value=res["과제명"]) # 과제명 (로드맵)
            sheet.cell(row=current_row, column=11, value=res["법적요구"]) # 법적요구 (로드맵)
            sheet.cell(row=current_row, column=12, value=int(res["시급성"])) # 시급성 (로드맵)
            sheet.cell(row=current_row, column=13, value=int(res["위험도"])) # 위험도 (로드맵)
            sheet.cell(row=current_row, column=14, value=res["예상예산"]) # 예상예산 (로드맵)
            sheet.cell(row=current_row, column=15, value=res["로드맵연도"]) # 로드맵연도 (로드맵)
            sheet.cell(row=current_row, column=16, value=res["비고"]) # 비고 (로드맵)
            
        # 미사용 기본 행들 지우거나 서식 유지 (새로 들어간 개수만큼 템플릿 하단 셀 정리 - 필요 시)
        wb.save(OUTPUT_PATH)
        print(f"성공: 로드맵 결과 엑셀 작성이 완료되었습니다: {OUTPUT_PATH}")
        return True
    except Exception as e:
        print(f"오류: 엑셀 파일 작성 중 에러 발생: {e}")
        return False

def main():
    print("="*60)
    print("      보안 로드맵 수립 에이전트 파이프라인 가동")
    print("="*60)
    
    # 1. 자사 보안 솔루션 로드
    sol_df = load_security_solutions()
    sol_context = build_markdown_context(sol_df)
    
    # 2. 고객사 체크리스트 파싱
    checklist_items = load_checklist()
    if not checklist_items:
        print("경고: 분석할 체크리스트 항목이 없습니다. 프로그램을 종료합니다.")
        return
        
    # 3. 매핑 및 Reasoning 루프
    print("[3/6] LLM / Rule 하이브리드 엔진 매핑 진행 중...")
    results = []
    
    # API 키 유무 체크하여 모드 출력
    api_key = os.getenv("GEMINI_API_KEY")
    if api_key and "your_gemini_api_key" not in api_key:
        print("-> [MODE] Gemini LLM + RAG 매핑 프로세스 구동")
    else:
        print("-> [MODE] 로컬 Rule-based 매핑 프로세스 구동 (Fallback)")
        
    for i, item in enumerate(checklist_items):
        print(f"   [{i+1}/{len(checklist_items)}] 항목 매핑 중: {item['항목명'][:15]}...")
        try:
            # 하이브리드 매핑 호출
            if api_key and "your_gemini_api_key" not in api_key:
                map_res = llm_based_mapping(item, sol_context)
            else:
                map_res = rule_based_mapping(item)
        except Exception as e:
            # API 키가 에러 날 경우 룰 매퍼 fallback
            map_res = rule_based_mapping(item)
            
        # 4. 컴플라이언스 스케줄러를 통한 연도 배정
        roadmap_year = calculate_roadmap_year(map_res)
        
        # 데이터 통합
        integrated_res = {
            "항목명": item["항목명"],
            "세부점검내용": item["세부점검내용"],
            "운영현황_증적": item["운영현황_증적"],
            "개선방안": item["개선방안"],
            "보안영역": map_res.get("보안영역", "기타 보안"),
            "과제명": map_res.get("과제명", "보안 대책 수립"),
            "법적요구": map_res.get("법적요구", "N/A"),
            "시급성": map_res.get("시급성", 3),
            "위험도": map_res.get("위험도", 3),
            "예상예산": map_res.get("예상예산", "₩25,000,000"),
            "로드맵연도": roadmap_year,
            "비고": map_res.get("비고", "")
        }
        results.append(integrated_res)
        
    # 5. 로드맵 템플릿에 최종 라이팅
    success = write_to_roadmap_template(results)
    
    # 6. 완료 상태 보고
    print("="*60)
    if success:
        print("      보안 솔루션 로드맵 수립 에이전트 완료")
        print(f"      결과물 저장 완료: {OUTPUT_PATH}")
    else:
        print("      로드맵 수립 도중 오류가 발생했습니다.")
    print("="*60)

if __name__ == "__main__":
    main()
