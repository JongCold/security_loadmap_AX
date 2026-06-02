import os
import re
import sys
import json
import collections
import math
from copy import copy
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import requests

# RAG 엔진 임포트
from rag_engine import get_rag_engine, initialize_rag

# 표준 출력 인코딩을 UTF-8로 설정 및 라인 버퍼링 강제
sys.stdout.reconfigure(encoding='utf-8', line_buffering=True)

# 경로 설정 - 프로젝트 디렉토리 기준 상대 경로 자동화
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SOLUTION_LIST_PATH = os.path.join(BASE_DIR, "자사 보안 솔루션 리스트 (1).xlsx")
ROADMAP_TEMPLATE_PATH = os.path.join(BASE_DIR, "{올해}년 KG그룹_{기업명} 정보보안감사_보안솔루션로드맵_{생성일}.xlsx")

# 로컬 Ollama 설정
OLLAMA_URL = "http://127.0.0.1:11434/api/generate"
OLLAMA_MODEL = "gemma2:2b"


def copy_cell_style(src_cell, dest_cell):
    """openpyxl용 셀 스타일 복사 헬퍼 함수"""
    if src_cell.has_style:
        dest_cell.font = copy(src_cell.font)
        dest_cell.fill = copy(src_cell.fill)
        dest_cell.border = copy(src_cell.border)
        dest_cell.alignment = copy(src_cell.alignment)
        dest_cell.number_format = src_cell.number_format


def get_merged_cell_value(sheet, row, start_col, end_col):
    """지정된 범위의 열을 순회하며 빈 값이 아닌 최초의 값을 반환"""
    for col in range(start_col, end_col + 1):
        val = sheet.cell(row=row, column=col).value
        if val is not None and str(val).strip() != "":
            return val
    return None


def is_red_font(cell):
    """셀의 글꼴 색상이 빨간색 계열인지 여부를 판별"""
    if not cell or not cell.font or not cell.font.color:
        return False
    color = cell.font.color
    
    # 1. RGB 값 직접 검사
    if color.type == 'rgb' and color.rgb:
        rgb_val = str(color.rgb).upper()
        if len(rgb_val) == 8:
            rgb_val = rgb_val[2:]  # Alpha 채널 제거
        if len(rgb_val) == 6:
            try:
                r = int(rgb_val[0:2], 16)
                g = int(rgb_val[2:4], 16)
                b = int(rgb_val[4:6], 16)
                if r > 180 and g < 110 and b < 110:
                    return True
            except ValueError:
                pass
                
    # 2. 테마 또는 다른 타입인 경우 빨간색 인덱스 검사 (보완용)
    if color.indexed == 10 or color.indexed == 2:
        return True
        
    return False


def is_red_fill(cell):
    """셀의 배경색이 빨간색 계열인지 여부를 판별"""
    if not cell or not cell.fill or not cell.fill.start_color:
        return False
    color = cell.fill.start_color
    
    # 1. RGB 값 직접 검사
    if color.type == 'rgb' and color.rgb:
        rgb_val = str(color.rgb).upper()
        if len(rgb_val) == 8:
            rgb_val = rgb_val[2:]  # Alpha 채널 제거
        if len(rgb_val) == 6:
            try:
                r = int(rgb_val[0:2], 16)
                g = int(rgb_val[2:4], 16)
                b = int(rgb_val[4:6], 16)
                if r > 180 and g < 110 and b < 110:
                    return True
            except ValueError:
                pass
                
    # 2. 테마 또는 다른 타입인 경우 빨간색 인덱스 검사 (보완용)
    if color.indexed == 10 or color.indexed == 2:
        return True
        
    return False


def find_best_matching_solution(item):
    """ChromaDB RAG 기반 자사 솔루션 최적 매핑"""
    try:
        rag = get_rag_engine()
        best_sol, similarity = rag.find_best_solution(item, similarity_threshold=0.20)

        if best_sol:
            return {
                "보안영역": best_sol.get("보안영역", ""),
                "솔루션구분": best_sol.get("솔루션구분", ""),
                "제조사명": best_sol.get("제조사명", ""),
                "제품명": best_sol.get("제품명", ""),
                "제품명기능설명": best_sol.get("제품명기능설명", "")
            }, similarity
        else:
            return None, similarity
    except Exception as e:
        print(f"[RAG] 매핑 에러: {e}", flush=True)
        return None, 0.0


def parse_red_cells_from_checklist(filepath):
    """상세체크리스트 파일에서 조치가 필요한 결함 사항(평가 'X') 및 기존 데이터 추출"""
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"체크리스트 파일을 찾을 수 없습니다: {filepath}")
        
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if "내부보안감사체크리스트" in name:
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[3]
        
    sheet = wb[sheet_name]
    
    # 헤더 동적 스캔 수립
    col_item_name, col_detail, col_eval, col_status, col_improv = 3, 4, 5, 6, 7
    col_area, col_project, col_law, col_urgency, col_risk, col_budget, col_year, col_note = 9, 10, 11, 12, 13, 14, 15, 16
    
    for r_hdr in [1, 2]:
        for c in range(1, min(25, sheet.max_column + 1)):
            val = sheet.cell(row=r_hdr, column=c).value
            if val and isinstance(val, str):
                val_clean = val.replace(" ", "").replace("\n", "").strip()
                if "항목명" in val_clean: col_item_name = c
                elif "세부점검내용" in val_clean: col_detail = c
                elif "평가" in val_clean: col_eval = c
                elif "운영현황" in val_clean or "보안감사" in val_clean: col_status = c
                elif "개선방안" in val_clean: col_improv = c
                elif "보안영역" in val_clean: col_area = c
                elif "과제명" in val_clean: col_project = c
                elif "법적요구" in val_clean: col_law = c
                elif "시급성" in val_clean: col_urgency = c
                elif "위험도" in val_clean: col_risk = c
                elif "예상예산" in val_clean: col_budget = c
                elif "로드맵연도" in val_clean or "추진연도" in val_clean: col_year = c
                elif "비고" in val_clean: col_note = c

    defect_items = []
    last_item_name = "미정"
    
    for r in range(3, sheet.max_row + 1):
        item_name_cell = sheet.cell(row=r, column=col_item_name)
        detail_cell = sheet.cell(row=r, column=col_detail)
        eval_cell = sheet.cell(row=r, column=col_eval)
        status_cell = sheet.cell(row=r, column=col_status)
        improv_cell = sheet.cell(row=r, column=col_improv)
        
        eval_val = str(eval_cell.value).strip().upper() if eval_cell.value else ""
        status_val = str(status_cell.value).strip() if status_cell.value else ""
        improv_val = str(improv_cell.value).strip() if improv_cell.value else ""
        
        if item_name_cell.value is not None and str(item_name_cell.value).strip() != "":
            last_item_name = str(item_name_cell.value).strip()
            
        if eval_val == "X":
            area_val = str(sheet.cell(row=r, column=col_area).value).strip() if sheet.cell(row=r, column=col_area).value else "기타 보안"
            project_val = str(sheet.cell(row=r, column=col_project).value).strip() if sheet.cell(row=r, column=col_project).value else "보안 솔루션 구축"
            law_val = str(sheet.cell(row=r, column=col_law).value).strip() if sheet.cell(row=r, column=col_law).value else "N/A"
            urgency_val = safe_int(sheet.cell(row=r, column=col_urgency).value, 3)
            risk_val = safe_int(sheet.cell(row=r, column=col_risk).value, 3)
            budget_val = str(sheet.cell(row=r, column=col_budget).value).strip() if sheet.cell(row=r, column=col_budget).value else "영업 문의가 필요한 영역 논의 필요."
            year_val = str(sheet.cell(row=r, column=col_year).value).strip() if sheet.cell(row=r, column=col_year).value else "2026년"
            note_val = str(sheet.cell(row=r, column=col_note).value).strip() if sheet.cell(row=r, column=col_note).value else ""

            rec_sol, rec_vendor = "N/A", "N/A"
            if note_val:
                match = re.search(r'\[추천\s*솔루션\]\s*(.*?)\s*-\s*([^\n\r]+)', note_val)
                if match:
                    rec_vendor = match.group(1).strip()
                    rec_sol = match.group(2).strip()

            defect_items.append({
                "row_idx": r, "항목명": last_item_name,
                "세부점검내용": str(detail_cell.value).strip() if detail_cell.value else "",
                "평가": eval_val, "운영현황_증적": status_val, "개선방안": improv_val,
                "보안영역": area_val, "과제명": project_val, "법적요구": law_val,
                "시급성": urgency_val, "위험도": risk_val, "예상예산": budget_val,
                "로드맵연도": year_val, "비고": note_val, "추천솔루션": rec_sol, "제조사": rec_vendor
            })
            
    return defect_items


def safe_int(val, default=3):
    """안전한 정수 형변환 헬퍼 함수"""
    try:
        if val is None: return default
        if isinstance(val, (int, float)): return int(val)
        nums = re.findall(r'\d+', str(val))
        return int(nums[0]) if nums else default
    except Exception:
        return default


def reformat_note(note_val, res_sol="N/A", res_vendor="N/A"):
    """
    기존 비고 포맷을 다음 사용자 요구 사항에 맞춰 변환하는 안전한 후처리 함수:
    [추천 솔루션] : {제품명}|{제조사}
    [선정 이유]
    1. 이유1
    2. 이유2
    """
    note_val = str(note_val).strip()
    
    # 1. 기존 값 파싱 또는 파라미터 값 기본값 설정
    sol_name = res_sol if res_sol else "N/A"
    vendor_name = res_vendor if res_vendor else "N/A"
    
    # note_val에서 [추천 솔루션] 값 파싱
    sol_match = re.search(r'\[추천\s*솔루션\]\s*:?\s*([^|\-\n]+?)\s*(?:-|\|)\s*([^\n]+)', note_val)
    if sol_match:
        p1 = sol_match.group(1).strip()
        p2 = sol_match.group(2).strip()
        vendors = ["ahnlab", "secui", "genians", "wins", "piolink", "igloosec", "somansa", "softcamp", "fasoo", "jiransoft", "kingsinformation", "estsoft", "cyberone", "kvine", "n/a"]
        if p1.lower() in vendors:
            vendor_name = p1
            sol_name = p2
        elif p2.lower() in vendors:
            vendor_name = p2
            sol_name = p1
        else:
            if res_sol and res_sol != "N/A":
                sol_name = res_sol
                vendor_name = res_vendor
            else:
                vendor_name = p1
                sol_name = p2
    else:
        # 단순 매치 (예: [추천 솔루션] AhnLab - V3 Office Security)
        sol_match_simple = re.search(r'\[추천\s*솔루션\]\s*([^\n\-]+?)\s*-\s*([^\n]+)', note_val)
        if sol_match_simple:
            vendor_name = sol_match_simple.group(1).strip()
            sol_name = sol_match_simple.group(2).strip()
        else:
            if "N/A" in note_val:
                sol_name = "N/A"
                vendor_name = "N/A"

    # N/A 대소문자 정리
    if sol_name.upper() == "NONE" or sol_name.upper() == "N/A" or not sol_name:
        sol_name = "N/A"
    if vendor_name.upper() == "NONE" or vendor_name.upper() == "N/A" or not vendor_name:
        vendor_name = "N/A"

    # 2. 선정 이유 파싱 및 재배열
    reason_text = ""
    reason_match = re.search(r'\[선정\s*이유\]\s*(.*)', note_val, re.DOTALL)
    if reason_match:
        reason_text = reason_match.group(1).strip()
    else:
        reason_text = re.sub(r'\[추천\s*솔루션\].*?(\n|$)', '', note_val).strip()
        
    reason_clean = re.sub(r'^\d+[\.\)\s\-]+', '', reason_text, flags=re.MULTILINE)
    reason_clean = re.sub(r'\s+', ' ', reason_clean).strip()
    
    # 문장 분할 (마침표 기준)
    sentences = []
    raw_sentences = re.split(r'\.\s*', reason_clean)
    for s in raw_sentences:
        s_clean = s.strip()
        if s_clean:
            if not s_clean.endswith('.'):
                s_clean += '.'
            if len(s_clean) > 3:
                sentences.append(s_clean)
                
    if not sentences:
        if sol_name != "N/A":
            sentences = [f"{sol_name} 도입을 통해 보안 정책을 강제화하고 외부 보안 위협에 대비한 통제 수준을 강화하기 위함입니다."]
        else:
            sentences = ["기술적 보안 솔루션 도입 대신 내부 보안 지침 수립 및 관리 감독 프로세스 강화를 권고합니다."]
            
    # 최종 재포맷 문자열 조립
    if sol_name == "N/A" and vendor_name == "N/A":
        rec_line = "[추천 솔루션] : N/A"
    else:
        rec_line = f"[추천 솔루션] : {sol_name}|{vendor_name}"
        
    lines = [rec_line, "[선정 이유]"]
    
    if len(sentences) == 1:
        lines.append(f"1. {sentences[0]}")
    else:
        for idx, sentence in enumerate(sentences[:3]):
            lines.append(f"{idx+1}. {sentence}")
            
    return "\n".join(lines)


def call_local_gemma(item, sol_context=None, model_name=OLLAMA_MODEL):
    """로컬 Ollama LLM 모델을 사용하여 솔루션 자동 추천 및 매핑 수행 (끊긴 프롬프트 및 완성 구조 복구)"""
    best_sol, score = find_best_matching_solution(item)
    print(f"[RAG] 매핑 결과: 제품={best_sol['제품명'] if best_sol else 'N/A'}, 유사도={score:.4f}", flush=True)
    
    rag = get_rag_engine()
    rag_context_md = rag.get_rag_context_markdown(item, top_k=3)
    
    if best_sol and score >= 0.20:
        rag_guideline = f"""
[RAG 매핑된 실제 자사 보안 솔루션 지정]
- 보안영역: {best_sol['보안영역']}
- 제품명: {best_sol['제품명']}
- 제조사명: {best_sol['제조사명']}
- 제품 기능 설명: {best_sol['제품명기능설명']}

[지시사항]
- "보안영역": 위 솔루션의 실제 보안영역("{best_sol['보안영역']}") 기재.
- "과제명": "{best_sol['제품명']} 도입 및 보안 통제 강화" 형태로 작성.
- "법적요구": 관련 법적 근거 규정 기재 (예: 개인정보 보호법 제29조(안전조치의무) 등. 없을 시 'N/A').
- "시급성": 1~5 사이의 정수값 (법적 요구가 존재하거나 시급한 경우 4~5로 설정).
- "위험도": 1~5 사이의 정수값.
- "예상예산": "₩30,000,000" 또는 솔루션 예산 규모 기재.
- "비고": 아래 서식을 정확히 준수하여 한국어로 작성. (선정이유는 문장 단위로 분할하여 번호를 매길 것)
  [추천 솔루션] : {best_sol['제품명']}|{best_sol['제조사명']}
  [선정 이유]
  1. 선정 이유 내용 1
  2. 선정 이유 내용 2
  3. 선정 이유 내용 3
"""
    else:
        rag_guideline = f"""
[RAG 매핑 결과: 적절한 도입 보안 솔루션 없음]
- "보안영역": "기타 보안" 또는 일반 관리 보안 영역 기재.
- "과제명": "보안 통제 절차 수립 및 규정 강화" 기재.
- "법적요구": 관련 법적 요건 명시 (없을 시 'N/A').
- "시급성": 1~5 사이의 정수값.
- "위험도": 1~5 사이의 정수값.
- "예상예산": "영업 문의가 필요한 영역 논의 필요." 기재.
- "비고": 아래 서식을 준수하여 작성.
  [추천 솔루션] : N/A
  [선정 이유]
  1. 기술적 장비 도입 대신 사내 보안 지침/절차 수립 및 수동 관리 감독 프로세스 강화를 권고합니다.
"""

    prompt = f"""
역할: 정보보안 CISO 컨설턴트 및 솔루션 아키텍트.
주어진 [고객사 결함사항]을 정밀 분석하고 최종 조치 로드맵 JSON을 구성하라.

{rag_guideline}

[고객사 결함사항]
- 진단 항목명: {item['항목명']}
- 점검 세부내용: {item['세부점검내용']}
- 실태 운영현황: {item['운영현황_증적']}
- 권고 개선방안: {item['개선방안']}

[출력 요구사항 및 JSON 스키마]
반드시 다른 설명 텍스트 없이 아래 명세 스펙의 키 이름을 가진 단일 JSON 객체 형태로만 대답하라.
{{
  "보안영역": "보안영역 명칭",
  "과제명": "과제명",
  "법적요구": "법적요구사항 또는 'N/A'",
  "시급성": 3,
  "위험도": 3,
  "예상예산": "예상예산 비용",
  "비고": "추천 솔루션 및 선정 이유 포맷"
}}
"""
    try:
        payload = {"model": model_name, "prompt": prompt, "stream": False, "format": "json"}
        res = requests.post(OLLAMA_URL, json=payload, timeout=90)
        if res.status_code == 200:
            response_json = res.json()
            res_data = json.loads(response_json.get("response", "{}"))
            
            # 후처리 보정 장치: 로컬 LLM이 반환하는 키 불일치 및 추천솔루션 누락 해결
            final_res = {}
            final_res["보안영역"] = res_data.get("보안영역", best_sol.get("보안영역", "기타 보안") if best_sol else "기타 보안")
            final_res["과제명"] = res_data.get("과제명", f"{best_sol['제품명']} 도입 및 보안 통제 강화" if best_sol else "보안 통제 절차 수립 및 규정 강화")
            final_res["법적요구"] = res_data.get("법적요구", "N/A")
            final_res["시급성"] = safe_int(res_data.get("시급성", 3), 3)
            final_res["위험도"] = safe_int(res_data.get("위험도", 3), 3)
            final_res["예상예산"] = res_data.get("예상예산", "영업 문의가 필요한 영역 논의 필요.")
            
            # 비고 후처리 자동 적용
            raw_note = res_data.get("비고", "")
            final_res["비고"] = reformat_note(
                raw_note, 
                res_sol=best_sol.get("제품명", "N/A") if best_sol else "N/A",
                res_vendor=best_sol.get("제조사명", "N/A") if best_sol else "N/A"
            )
            
            if best_sol:
                final_res["추천솔루션"] = best_sol.get("제품명", "N/A")
                final_res["제조사"] = best_sol.get("제조사명", "N/A")
            else:
                final_res["추천솔루션"] = "N/A"
                final_res["제조사"] = "N/A"
                
            return final_res
    except Exception as e:
        print(f"[LLM 호출 에러] {e}", flush=True)
        
    fallback = {
        "보안영역": best_sol.get("보안영역", "기타 보안") if best_sol else "기타 보안",
        "과제명": f"{best_sol['제품명']} 도입 및 보안 통제 강화" if best_sol else "보안 통제 절차 수립 및 규정 강화",
        "법적요구": "N/A",
        "시급성": 3,
        "위험도": 3,
        "예상예산": "영업 문의가 필요한 영역 논의 필요.",
        "비고": reformat_note("", best_sol.get("제품명", "N/A") if best_sol else "N/A", best_sol.get("제조사명", "N/A") if best_sol else "N/A"),
        "추천솔루션": best_sol.get("제품명", "N/A") if best_sol else "N/A",
        "제조사": best_sol.get("제조사명", "N/A") if best_sol else "N/A"
    }
    return fallback


def generate_roadmap_excel(results, company_name, output_filepath, env_filepath=None, asset_filepath=None):
    """원본 로드맵 템플릿 서식을 유지하며 매핑 결과 및 사전환경조사서, 자산목록 데이터를 통합하여 작성"""
    
    def count_lines(text, max_chars):
        if not text:
            return 1
        lines = str(text).split('\n')
        total = 0
        for line in lines:
            v_len = 0
            for ch in line:
                if ord(ch) > 127: # 한글 등 멀티바이트
                    v_len += 2
                else:
                    v_len += 1
            total += max(1, math.ceil(v_len / max_chars))
        return total

    if not os.path.exists(ROADMAP_TEMPLATE_PATH):
        raise FileNotFoundError(f"템플릿 엑셀 파일을 찾을 수 없습니다: {ROADMAP_TEMPLATE_PATH}")
        
    wb = openpyxl.load_workbook(ROADMAP_TEMPLATE_PATH)
    sheet_name = None
    for name in wb.sheetnames:
        if "01_KG그룹" in name or "01_" in name:
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[1]
        
    sheet = wb[sheet_name]
    
    # 0. 템플릿 레이아웃 정보 동적 분석
    asset_title_row, roadmap_title_row = 11, 23
    for r in range(1, sheet.max_row + 1):
        val = sheet.cell(row=r, column=1).value
        if val and isinstance(val, str):
            val_clean = val.replace(" ", "").replace("\n", "").strip()
            if "보안솔루션세부현황" in val_clean: asset_title_row = r
            elif "보안솔루션로드맵" in val_clean: roadmap_title_row = r
                
    asset_data_start = asset_title_row + 2
    asset_data_end = roadmap_title_row - 1
    default_asset_rows = asset_data_end - asset_data_start + 1
    
    roadmap_data_start = roadmap_title_row + 3
    roadmap_data_end = sheet.max_row
    if roadmap_data_end < roadmap_data_start:
        roadmap_data_end = roadmap_data_start + 4
    default_roadmap_rows = roadmap_data_end - roadmap_data_start + 1

    original_heights = {r: sheet.row_dimensions[r].height for r in range(1, sheet.max_row + 1) if r in sheet.row_dimensions}
    original_merges = list(sheet.merged_cells.ranges)
    for m_range in original_merges:
        sheet.unmerge_cells(start_row=m_range.min_row, start_column=m_range.min_col, end_row=m_range.max_row, end_column=m_range.max_col)
    
    # 1. 사전환경조사서 반영
    if env_filepath and os.path.exists(env_filepath):
        try:
            env_wb = openpyxl.load_workbook(env_filepath, data_only=True)
            env_sheet = None
            for name in env_wb.sheetnames:
                if "KG그룹" in name or "가족사명" in name:
                    env_sheet = env_wb[name]
                    break
            if not env_sheet:
                for name in env_wb.sheetnames:
                    if "환경조사서" in name:
                        env_sheet = env_wb[name]
                        break
            if not env_sheet:
                env_sheet = env_wb.worksheets[2] if len(env_wb.worksheets) > 2 else env_wb.worksheets[1] if len(env_wb.worksheets) > 1 else env_wb.active
                
            # 신청 업체 정보의 병합 셀들을 고려하여 범위 내에서 파싱
            # 기업명 : B ~ F열 (2~6열)
            company_name_val = get_merged_cell_value(env_sheet, 5, 2, 6)
            # 사업 형태 : B ~ F열 (2~6열)
            business_type_val = get_merged_cell_value(env_sheet, 6, 2, 6)
            # 임직원 수 : B ~ F열 (2~6열)
            employee_count_val = get_merged_cell_value(env_sheet, 7, 2, 6)
            # 주요 적용 법률 : B ~ F열 (2~6열)
            law_val = get_merged_cell_value(env_sheet, 8, 2, 6)

            # 사업분야 : H ~ K열 (8~11열)
            business_area_val = get_merged_cell_value(env_sheet, 5, 8, 11)
            # 정보통신 사업자 여부 : H ~ K열 (8~11열)
            it_operator_val = get_merged_cell_value(env_sheet, 6, 8, 11)
            # 전년도 매출 : H ~ K열 (8~11열)
            sales_val = get_merged_cell_value(env_sheet, 7, 8, 11)
            # 감독기관(유관 정부 기관) : H ~ K열 (8~11열)
            governing_val = get_merged_cell_value(env_sheet, 8, 8, 11)

            sheet.cell(row=5, column=2, value=str(company_name_val).strip() if company_name_val else "")
            sheet.cell(row=5, column=9, value=str(business_area_val).strip() if business_area_val else "")
            sheet.cell(row=6, column=2, value=str(business_type_val).strip() if business_type_val else "")
            sheet.cell(row=6, column=9, value=str(it_operator_val).strip() if it_operator_val else "")
            
            if employee_count_val is not None:
                sheet.cell(row=7, column=2, value=employee_count_val)
            else:
                sheet.cell(row=7, column=2, value="")
                
            sheet.cell(row=7, column=9, value=str(sales_val).strip() if sales_val else "")
            sheet.cell(row=8, column=2, value=str(law_val).strip() if law_val else "")
            sheet.cell(row=8, column=9, value=str(governing_val).strip() if governing_val else "")
        except Exception as env_err:
            print(f"[엑셀 병합] 사전환경조사서 파싱 에러: {env_err}", flush=True)

    # 1.5. 12행 헤더 영역 리포맷 (기밀성, 무결성, 가용성, 합계, 등급, 비고 헤더 적용)
    sheet.cell(row=12, column=9, value="기밀성")
    sheet.cell(row=12, column=10, value="무결성")
    sheet.cell(row=12, column=11, value="가용성")
    sheet.cell(row=12, column=12, value="합계")
    sheet.cell(row=12, column=13, value="등급")
    sheet.cell(row=12, column=14, value="비고")
    
    # 12행 헤더 정렬 및 서식을 기존 8열(담당부서) 헤더 서식으로부터 이식
    for h_col in range(9, 14):
        copy_cell_style(sheet.cell(row=12, column=8), sheet.cell(row=12, column=h_col))
        sheet.cell(row=12, column=h_col).alignment = Alignment(horizontal="center", vertical="center")
    
    # 12행 비고 영역(14~16열) 전체에 스타일 및 가운데 정렬 복사 적용 후 병합
    for h_col in range(14, 17):
        copy_cell_style(sheet.cell(row=12, column=8), sheet.cell(row=12, column=h_col))
        sheet.cell(row=12, column=h_col).alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells(start_row=12, start_column=14, end_row=12, end_column=16)

    # 2. 자산목록 데이터 동적 이식
    diff = 0
    N = 0
    thin_side = Side(style='thin', color='D3D3D3')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    if asset_filepath and os.path.exists(asset_filepath):
        try:
            asset_wb = openpyxl.load_workbook(asset_filepath, data_only=True)
            asset_sheet = None
            for name in asset_wb.sheetnames:
                if "정보보호시스템" in name.replace(" ", ""):
                    asset_sheet = asset_wb[name]
                    break
            if not asset_sheet: 
                asset_sheet = asset_wb.active
                
            asset_rows = []
            for r in range(9, asset_sheet.max_row + 1):
                no_val = asset_sheet.cell(row=r, column=2).value
                if no_val is None: 
                    continue
                
                # 자산 정보 동적 정리 추출
                gubun = str(asset_sheet.cell(row=r, column=4).value or "").strip()
                ip = str(asset_sheet.cell(row=r, column=5).value or "").strip()
                vendor = str(asset_sheet.cell(row=r, column=7).value or "").strip()
                func = str(asset_sheet.cell(row=r, column=8).value or "").strip()
                
                if not gubun and not ip and not vendor: 
                    continue
                
                # 운영 미가동(비활성) 자산 감지
                is_inactive = False
                for col_idx in range(1, 16):
                    if is_red_fill(asset_sheet.cell(row=r, column=col_idx)):
                        is_inactive = True
                        break
                
                # 각 셀의 배경색(fill) 추출
                fills = {
                    "구분": asset_sheet.cell(row=r, column=4).fill,
                    "IP주소": asset_sheet.cell(row=r, column=5).fill,
                    "벤더": asset_sheet.cell(row=r, column=7).fill,
                    "기능": asset_sheet.cell(row=r, column=8).fill,
                    "자산위치": asset_sheet.cell(row=r, column=9).fill,
                    "담당자": asset_sheet.cell(row=r, column=10).fill,
                    "담당부서": asset_sheet.cell(row=r, column=12).fill,
                    "기밀성": asset_sheet.cell(row=r, column=13).fill,
                    "무결성": asset_sheet.cell(row=r, column=14).fill,
                    "가용성": asset_sheet.cell(row=r, column=15).fill,
                    "합계": asset_sheet.cell(row=r, column=16).fill,
                    "등급": asset_sheet.cell(row=r, column=17).fill,
                }
                
                asset_rows.append({
                    "구분": gubun, 
                    "IP주소": ip, 
                    "벤더": vendor, 
                    "기능": func,
                    "자산위치": str(asset_sheet.cell(row=r, column=9).value or "").strip(),
                    "담당자": str(asset_sheet.cell(row=r, column=10).value or "").strip(),
                    "담당부서": str(asset_sheet.cell(row=r, column=12).value or "").strip(),
                    "기밀성": safe_int(asset_sheet.cell(row=r, column=13).value, 3),
                    "무결성": safe_int(asset_sheet.cell(row=r, column=14).value, 3),
                    "가용성": safe_int(asset_sheet.cell(row=r, column=15).value, 3),
                    "등급": str(asset_sheet.cell(row=r, column=17).value or "M").strip(),
                    "비고": "운영하고 있지 않는 자산" if is_inactive else "",
                    "fills": fills
                })
                
            N = len(asset_rows)
            if N > 0:
                diff = N - default_asset_rows
                if diff > 0:
                    sheet.insert_rows(roadmap_title_row, amount=diff)
                    for r in range(roadmap_title_row, roadmap_title_row + diff):
                        sheet.row_dimensions[r].height = sheet.row_dimensions[asset_data_start].height
                        for col_idx in range(1, 17):
                            copy_cell_style(sheet.cell(row=asset_data_start, column=col_idx), sheet.cell(row=r, column=col_idx))
                elif diff < 0:
                    sheet.delete_rows(asset_data_start + N, amount=-diff)
                    
                # 자산 데이터 셀 기입 및 포맷 보존
                for i, asset in enumerate(asset_rows):
                    r = asset_data_start + i
                    
                    # 자산목록의 기능(5열), 비고(14열) 등을 분석하여 동적 행 높이 결정
                    a_func = asset.get("기능", "")
                    a_note = asset.get("비고", "")
                    h_func = count_lines(a_func, 30)
                    h_note = count_lines(a_note, 30)
                    max_asset_lines = max(h_func, h_note)
                    sheet.row_dimensions[r].height = max(28, max_asset_lines * 16 + 8)

                    sheet.cell(row=r, column=1, value=i + 1)
                    
                    cols_map = {
                        2: ("구분", asset["구분"]),
                        3: ("IP주소", asset["IP주소"]),
                        4: ("벤더", asset["벤더"]),
                        5: ("기능", asset["기능"]),
                        6: ("자산위치", asset["자산위치"]),
                        7: ("담당자", asset["담당자"]),
                        8: ("담당부서", asset["담당부서"]),
                        9: ("기밀성", asset["기밀성"]),
                        10: ("무결성", asset["무결성"]),
                        11: ("가용성", asset["가용성"]),
                        12: ("합계", f"=SUM(I{r}:K{r})"),
                        13: ("등급", asset["등급"]),
                        14: ("비고", asset["비고"])
                    }
                    
                    # 1열부터 16열까지 모든 열에 대해 스타일 복사 및 테두리 초기화
                    for col_idx in range(1, 17):
                        copy_cell_style(sheet.cell(row=asset_data_start, column=col_idx), sheet.cell(row=r, column=col_idx))
                        sheet.cell(row=r, column=col_idx).border = thin_border
                    
                    # 데이터 채우기 및 원본 셀 배경색 이식
                    for col_idx, (key, val) in cols_map.items():
                        cell = sheet.cell(row=r, column=col_idx, value=val)
                        
                        # 원본 배경색 이식
                        fl = asset["fills"].get(key)
                        if fl and fl.fill_type: 
                            cell.fill = copy(fl)
                            
                    # N열(14)~P열(16) 병합 영역 배경색/테두리 통일화 후 병합 설정 (비고 영역)
                    fill_to_apply = sheet.cell(row=r, column=14).fill
                    for col_b in range(14, 17):
                        sheet.cell(row=r, column=col_b).fill = copy(fill_to_apply)
                        sheet.cell(row=r, column=col_b).border = thin_border
                    sheet.merge_cells(start_row=r, start_column=14, end_row=r, end_column=16)
                    
                    # 정렬 맞춤 설정
                    for c_idx in range(1, 17):
                        if c_idx in [1, 2, 3, 6, 7, 8]:
                            sheet.cell(row=r, column=c_idx).alignment = Alignment(horizontal="center", vertical="center")
                        elif c_idx in [9, 10, 11, 12, 13, 14, 15, 16]:
                            sheet.cell(row=r, column=c_idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        except Exception as asset_err:
            print(f"[엑셀 병합] 자산목록 처리 오류: {asset_err}", flush=True)
 
    # 3. 로드맵 매핑 결과 데이터 기입
    start_row = roadmap_data_start + diff
    style_source_row = roadmap_data_start + diff
    M = len(results)
    
    roadmap_diff = M - default_roadmap_rows
    if roadmap_diff > 0:
        sheet.insert_rows(start_row + default_roadmap_rows, amount=roadmap_diff)
        for r in range(start_row + default_roadmap_rows, start_row + default_roadmap_rows + roadmap_diff):
            sheet.row_dimensions[r].height = sheet.row_dimensions[style_source_row].height
            for col_idx in range(1, 17):
                copy_cell_style(sheet.cell(row=style_source_row, column=col_idx), sheet.cell(row=r, column=col_idx))
    elif roadmap_diff < 0:
        sheet.delete_rows(start_row + M, amount=-roadmap_diff)
    
    for idx, res in enumerate(results):
        current_row = start_row + idx
        for col_idx in range(1, 17):
            copy_cell_style(sheet.cell(row=style_source_row, column=col_idx), sheet.cell(row=current_row, column=col_idx))
            
        sheet.cell(row=current_row, column=1, value=idx + 1)
        sheet.cell(row=current_row, column=2, value=res.get("항목명", ""))
        sheet.cell(row=current_row, column=3, value=res.get("세부점검내용", ""))
        sheet.cell(row=current_row, column=5, value=res.get("운영현황_증적", ""))
        sheet.cell(row=current_row, column=7, value=res.get("개선방안", ""))
        sheet.cell(row=current_row, column=9, value=res.get("보안영역", ""))
        sheet.cell(row=current_row, column=10, value=res.get("과제명", ""))
        sheet.cell(row=current_row, column=11, value=res.get("법적요구", "N/A"))
        sheet.cell(row=current_row, column=12, value=safe_int(res.get("시급성", 3)))
        sheet.cell(row=current_row, column=13, value=safe_int(res.get("위험도", 3)))
        sheet.cell(row=current_row, column=14, value=res.get("예상예산", "비용 발생 안 함"))
        sheet.cell(row=current_row, column=15, value=res.get("로드맵연도", "2026년"))
 
        # 비고 필드 신규 가시성 포맷 후처리 강제 적용
        note_val = res.get("비고", "")
        rec_sol = res.get("추천솔루션", "N/A")
        rec_vendor = res.get("제조사", "N/A")
        reformatted_note = reformat_note(note_val, res_sol=rec_sol, res_vendor=rec_vendor)
        sheet.cell(row=current_row, column=16, value=reformatted_note)

        # 개선방안(7, 8열) 폰트 색상을 빨간색(FF0000)으로 설정
        for c_idx in [7, 8]:
            cell_improv = sheet.cell(row=current_row, column=c_idx)
            if cell_improv.font:
                cell_improv.font = Font(
                    name=cell_improv.font.name,
                    size=cell_improv.font.size,
                    bold=cell_improv.font.bold,
                    italic=cell_improv.font.italic,
                    color="FF0000"
                )

        # 비고(16열) 폰트 색상을 검정색(000000)으로 설정
        cell_note = sheet.cell(row=current_row, column=16)
        if cell_note.font:
            cell_note.font = Font(
                name=cell_note.font.name,
                size=cell_note.font.size,
                bold=cell_note.font.bold,
                italic=cell_note.font.italic,
                color="000000"
            )

        # 정렬 설정 및 wrap_text 강제 적용
        alignments = {
            1: Alignment(horizontal="center", vertical="center"),
            2: Alignment(horizontal="center", vertical="center", wrap_text=True),
            3: Alignment(horizontal="left", vertical="center", wrap_text=True),
            4: Alignment(horizontal="left", vertical="center", wrap_text=True),
            5: Alignment(horizontal="left", vertical="center", wrap_text=True),
            6: Alignment(horizontal="left", vertical="center", wrap_text=True),
            7: Alignment(horizontal="left", vertical="center", wrap_text=True),
            8: Alignment(horizontal="left", vertical="center", wrap_text=True),
            9: Alignment(horizontal="center", vertical="center", wrap_text=True),
            10: Alignment(horizontal="left", vertical="center", wrap_text=True),
            11: Alignment(horizontal="center", vertical="center", wrap_text=True),
            12: Alignment(horizontal="center", vertical="center"),
            13: Alignment(horizontal="center", vertical="center"),
            14: Alignment(horizontal="center", vertical="center", wrap_text=True),
            15: Alignment(horizontal="center", vertical="center"),
            16: Alignment(horizontal="left", vertical="center", wrap_text=True),
        }
        for col_idx in range(1, 17):
            cell = sheet.cell(row=current_row, column=col_idx)
            if col_idx in alignments:
                cell.alignment = alignments[col_idx]

        # 행 높이 동적 계산 적용 (텍스트 길이에 대응)
        h_detail = count_lines(res.get("세부점검내용", ""), 40)
        h_status = count_lines(res.get("운영현황_증적", ""), 45)
        h_improv = count_lines(res.get("개선방안", ""), 50)
        h_note = count_lines(reformatted_note, 40)
        
        max_lines = max(h_detail, h_status, h_improv, h_note)
        calculated_height = max(35, max_lines * 16 + 12)
        sheet.row_dimensions[current_row].height = calculated_height
 
    # 4. 통합 병합 일괄 복구 적용
    for m_range in original_merges:
        min_row, max_row, min_col, max_col = m_range.min_row, m_range.max_row, m_range.min_col, m_range.max_col
        if min_row <= 12:
            if min_row == 12: 
                continue # 12행 헤더 병합은 수동으로 N12:P12 병합해 주므로 복원 대상에서 제외
            sheet.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        elif asset_data_start <= min_row <= asset_data_end:
            continue
        elif asset_data_end < min_row < roadmap_data_start:
            sheet.merge_cells(start_row=min_row + diff, start_column=min_col, end_row=max_row + diff, end_column=max_col)
        elif min_row >= roadmap_data_start:
            if min_row <= roadmap_data_end: 
                continue
            sheet.merge_cells(start_row=min_row + diff + roadmap_diff, start_column=min_col, end_row=max_row + diff + roadmap_diff, end_column=max_col)
 
    # (2) 결과 셀 신규 가로 병합 적용
    for idx in range(M):
        current_row = start_row + idx
        sheet.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
        sheet.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
        sheet.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=8)
 
    # 5. 워크시트 고객사명 동적 수정 반영
    for name in wb.sheetnames:
        if "01_KG그룹" in name or "01_" in name:
            wb[name].title = f"01_KG그룹({company_name})"
            break
 
    # 6. 열 너비 일괄 최적화 (가독성 향상)
    column_widths = {
        'A': 6,   # No
        'B': 18,  # 항목명
        'C': 22,  # 세부점검내용 (병합되므로 C+D)
        'D': 22,
        'E': 22,  # 운영현황 (병합되므로 E+F)
        'F': 22,
        'G': 25,  # 개선방안 (병합되므로 G+H)
        'H': 25,
        'I': 15,  # 보안영역
        'J': 28,  # 과제명
        'K': 22,  # 법적요구
        'L': 10,  # 시급성
        'M': 10,  # 위험도
        'N': 18,  # 예상예산
        'O': 12,  # 로드맵연도
        'P': 45,  # 비고 (추천 솔루션 : 제품명|제조사)
    }
    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

    wb.save(output_filepath)
    return output_filepath


def calculate_roadmap_year(mapping_res):
    """시급성, 위험도, 법적요구 기반 로드맵 연도(2026 ~ 2028) 산정 스케줄러"""
    urgency = int(mapping_res.get("시급성", 3))
    risk = int(mapping_res.get("위험도", 3))
    law = mapping_res.get("법적요구", "N/A")
    
    score = urgency * risk
    
    # 1. 법적 요구가 존재하면 1년차(2026년) 강제 최우선 배치
    if law and law != "N/A":
        return "2026년"
        
    # 2. 스코어 매트릭스에 따른 분배
    if score >= 16:
        return "2026년"
    elif score >= 8:
        return "2027년"
    else:
        return "2028년"