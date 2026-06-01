import os
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
template_file = os.path.join(BASE_DIR, "2026년 KG그룹_제로인 정보보안감사_보안솔루션로드맵_v1.2.xlsx")

wb = openpyxl.load_workbook(template_file, data_only=True)
# 01_KG그룹... 시트 찾기
sheet_name = None
for name in wb.sheetnames:
    if "01_KG그룹" in name or "01_" in name:
        sheet_name = name
        break
if not sheet_name:
    sheet_name = wb.sheetnames[1]

sheet = wb[sheet_name]
print(f"Sheet Name: {sheet.title}")
print("Row 12 headers (col 1 to 20):")
for col in range(1, 21):
    val = sheet.cell(row=12, column=col).value
    print(f"Col {col}: {val}")

# Row 13도 병합 구조가 있는지 확인하기 위해 출력
print("\nRow 13 details:")
for col in range(1, 21):
    cell = sheet.cell(row=13, column=col)
    print(f"Col {col}: value={cell.value}, cell_type={type(cell)}")

# 병합 셀 목록 조회
print("\nMerged cells:")
for merge in list(sheet.merged_cells.ranges):
    if merge.min_row <= 15:
        print(merge)
