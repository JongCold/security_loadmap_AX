import os
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
template_file = os.path.join(BASE_DIR, "2026년 KG그룹_제로인 정보보안감사_보안솔루션로드맵_v1.2.xlsx")

wb = openpyxl.load_workbook(template_file, data_only=True)
sheet_name = None
for name in wb.sheetnames:
    if "01_KG그룹" in name or "01_" in name:
        sheet_name = name
        break
if not sheet_name:
    sheet_name = wb.sheetnames[1]

sheet = wb[sheet_name]
print(f"Sheet Name: {sheet.title}")

# 26~35행 출력
for r in range(26, 36):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 18)]
    print(f"Row {r}: {row_vals}")
