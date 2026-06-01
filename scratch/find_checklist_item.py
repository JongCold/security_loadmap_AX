import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\USER\Downloads\Security loadmap_Auto-20260529T052519Z-3-001\Security loadmap_Auto"
checklist_path = os.path.join(base_dir, "uploads", "20260601_153104_2026년_KG이니시스_내부보안점검_상세체크리스트_v1.4_20260422.xlsx")

def find_item_in_checklist():
    print("=== 분석: 체크리스트 내 10.6 항목 탐색 ===")
    if not os.path.exists(checklist_path):
        print(f"체크리스트 파일이 존재하지 않습니다: {checklist_path}")
        return
    wb = openpyxl.load_workbook(checklist_path, data_only=True)
    
    sheet_name = None
    for name in wb.sheetnames:
        if "내부보안감사체크리스트" in name:
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[3]
    sheet = wb[sheet_name]
    print(f"선택된 체크리스트 시트: {sheet.title}")
    
    found = False
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=r, column=c).value
            if val and "10.6" in str(val):
                print(f"Row {r}, Col {c}에서 '10.6' 발견!")
                # 그 행의 전체 컬럼 값을 출력
                row_vals = [sheet.cell(row=r, column=col).value for col in range(1, 18)]
                print(f"  Row {r} 전체 데이터: {row_vals}")
                found = True
                break
                
    if not found:
        print("체크리스트 전체에서 '10.6'을 찾지 못했습니다.")

if __name__ == "__main__":
    find_item_in_checklist()
