import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
env_file = os.path.join(BASE_DIR, "uploads", "20260601_153113_2025년_KG그룹_정보보안감사_사전환경조사서_KG이니시스_2.xlsx")

if not os.path.exists(env_file):
    print("Env file not found.")
    exit(1)

wb = openpyxl.load_workbook(env_file, data_only=True)
print("Sheet Names:", wb.sheetnames)

# KG그룹 시트 이름 찾기
sheet_name = None
for name in wb.sheetnames:
    if "KG그룹" in name:
        sheet_name = name
        break
if not sheet_name:
    sheet_name = wb.worksheets[2].title if len(wb.worksheets) > 2 else wb.active.title

sheet = wb[sheet_name]
print(f"Reading Sheet: {sheet.title}")

for r in range(1, 15):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 11)]
    print(f"Row {r}: {row_vals}")
