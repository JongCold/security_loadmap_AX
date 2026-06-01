import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
checklist_file = os.path.join(BASE_DIR, "uploads", "20260601_153104_2026년_KG이니시스_내부보안점검_상세체크리스트_v1.4_20260422.xlsx")

wb = openpyxl.load_workbook(checklist_file, data_only=True)
sheet_name = None
for name in wb.sheetnames:
    if "내부보안감사체크리스트" in name:
        sheet_name = name
        break
if not sheet_name:
    sheet_name = wb.sheetnames[3]

sheet = wb[sheet_name]
print(f"Checklist Sheet: {sheet.title}")

for r in range(1, 6):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(12, sheet.max_column + 1))]
    print(f"Row {r}: {row_vals}")
