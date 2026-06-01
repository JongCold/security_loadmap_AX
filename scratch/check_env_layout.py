import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
env_file = os.path.join(BASE_DIR, "KG그룹 정보보안감사_사전환경조사서_양식.xlsx")

wb = openpyxl.load_workbook(env_file, data_only=True)
print("Sheet Names:", wb.sheetnames)

sheet = wb.worksheets[1] # 2번째 시트 (index 1)
print(f"Sheet Name: {sheet.title}")

for r in range(1, 16):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 16)]
    print(f"Row {r}: {row_vals}")

# 병합 범위
print("\n--- Merged Ranges ---")
for r_range in list(sheet.merged_cells.ranges):
    print(r_range)
