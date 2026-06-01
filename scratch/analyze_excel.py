import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\USER\Downloads\Security loadmap_Auto-20260529T052519Z-3-001\Security loadmap_Auto"
test_output_path = os.path.join(base_dir, "scratch", "test_output.xlsx")

def analyze_output_merges():
    print("=== 분석: 생성된 테스트 엑셀 병합 무결성 검증 ===")
    if not os.path.exists(test_output_path):
        print(f"테스트 출력 파일이 존재하지 않습니다: {test_output_path}")
        return
        
    wb = openpyxl.load_workbook(test_output_path, data_only=True)
    sheet = wb.worksheets[1] # 두 번째 시트
    print(f"시트 이름: {sheet.title}")
    print(f"최대 행 수: {sheet.max_row}")
    
    merged_ranges = list(sheet.merged_cells.ranges)
    print(f"전체 병합 범위 개수: {len(merged_ranges)}")
    
    # 18행과 19행 (로드맵 데이터 행)의 병합 상태 체크
    print("\n[로드맵 데이터 행(18, 19행) 병합 확인]")
    for r in [18, 19]:
        for c in range(1, 17):
            is_merged = False
            for m_range in merged_ranges:
                if m_range.min_row <= r <= m_range.max_row and m_range.min_col <= c <= m_range.max_col:
                    is_merged = True
                    break
            
            # C-D, E-F, G-H가 정상 병합되었는지 검사
            if c in [3, 4]: # C, D
                print(f"  Row {r}, Col {c} ({openpyxl.utils.get_column_letter(c)}{r}) 병합 여부: {is_merged}")
            elif c in [5, 6]: # E, F
                print(f"  Row {r}, Col {c} ({openpyxl.utils.get_column_letter(c)}{r}) 병합 여부: {is_merged}")
            elif c in [7, 8]: # G, H
                print(f"  Row {r}, Col {c} ({openpyxl.utils.get_column_letter(c)}{r}) 병합 여부: {is_merged}")
            
    # 전체 시트에서 의도하지 않게 꼬인 병합 범위가 있는지 목록을 간략히 출력
    print("\n[시트 전체 병합 범위 목록]")
    for m_range in sorted(merged_ranges, key=lambda x: x.min_row):
        print(f"  Merged: {m_range.coord}")

if __name__ == "__main__":
    analyze_output_merges()
