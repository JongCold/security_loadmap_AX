import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
xlsx_file = os.path.join(BASE_DIR, "2026년 KG그룹_KG이니시스 정보보안감사_보안솔루션로드맵_20260601 (1).xlsx")

if not os.path.exists(xlsx_file):
    print("파일이 존재하지 않습니다:", xlsx_file)
    sys.exit(1)

wb = openpyxl.load_workbook(xlsx_file, data_only=True)
print("Sheet Names:", wb.sheetnames)

sheet = wb.active
for name in wb.sheetnames:
    if "01_" in name or "KG그룹" in name:
        sheet = wb[name]
        break

print(f"Sheet Title: {sheet.title}")

# 11~20행 출력
for r in range(11, 22):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 18)]
    print(f"Row {r}: {row_vals}")

print("\n--- Merged Ranges ---")
for r_range in list(sheet.merged_cells.ranges):
    if r_range.min_row <= 25:
        print(r_range)
