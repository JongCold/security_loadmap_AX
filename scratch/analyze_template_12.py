import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
template_path = os.path.join(BASE_DIR, "2026년 KG그룹_제로인 정보보안감사_보안솔루션로드맵_v1.2.xlsx")

wb = openpyxl.load_workbook(template_path, data_only=True)
sheet = wb.active
for name in wb.sheetnames:
    if "01_" in name or "KG그룹" in name:
        sheet = wb[name]
        break

print(f"Sheet Title: {sheet.title}")
print("Row 12:", [sheet.cell(row=12, column=c).value for c in range(1, 18)])
print("Row 13:", [sheet.cell(row=13, column=c).value for c in range(1, 18)])

print("\nMerged ranges containing row 12 or 13:")
for r_range in list(sheet.merged_cells.ranges):
    if r_range.min_row <= 13 <= r_range.max_row or r_range.min_row <= 12 <= r_range.max_row:
        print(r_range)
