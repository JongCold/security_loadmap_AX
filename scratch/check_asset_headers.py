import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
asset_file = os.path.join(BASE_DIR, "uploads", "20260601_153119_KG그룹KG이니시스자산목록_및_중요도_평가서_20260601.xlsx")

wb = openpyxl.load_workbook(asset_file, data_only=True)
print("Sheet Names:", wb.sheetnames)

sheet_name = None
for name in wb.sheetnames:
    if "4.정보보호시스템" in name:
        sheet_name = name
        break
if not sheet_name:
    for name in wb.sheetnames:
        if "정보보호시스템" in name:
            sheet_name = name
            break
if not sheet_name:
    sheet_name = wb.active.title

sheet = wb[sheet_name]
print(f"Reading Sheet: {sheet.title}")

for r in range(1, 13):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 17)]
    print(f"Row {r:02d}: {row_vals}")
