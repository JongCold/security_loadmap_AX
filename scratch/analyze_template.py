import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

template_path = r"c:\Users\USER\Downloads\Security loadmap_Auto-20260529T052519Z-3-001\Security loadmap_Auto\2026년 KG그룹_제로인 정보보안감사_보안솔루션로드맵_v1.2.xlsx"

def analyze_template():
    print("=== 분석: 템플릿 파일 ===")
    if not os.path.exists(template_path):
        print(f"템플릿 파일이 존재하지 않습니다: {template_path}")
        return
    wb = openpyxl.load_workbook(template_path, data_only=True)
    sheet = wb.active
    for name in wb.sheetnames:
        if "01_" in name or "KG그룹" in name:
            sheet = wb[name]
            break
            
    print(f"템플릿 시트 이름: {sheet.title}")
    merged_ranges = list(sheet.merged_cells.ranges)
    print(f"전체 병합 범위 개수: {len(merged_ranges)}")
    
    # 모든 병합 범위를 행 번호 순으로 정렬하여 출력
    print("\n[전체 병합 범위 목록]")
    for m_range in sorted(merged_ranges, key=lambda x: x.min_row):
        print(f"  Row {m_range.min_row} ~ {m_range.max_row}, Col {m_range.min_col} ~ {m_range.max_col} ({m_range.coord})")

if __name__ == "__main__":
    analyze_template()
