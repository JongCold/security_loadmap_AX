import openpyxl
import os
import sys

# utf-8 설정
sys.stdout.reconfigure(encoding='utf-8')

filepath = "2026년 KG제로인_내부보안점검_상세체크리스트_테스트 파일.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=True)

sheet_name = None
for name in wb.sheetnames:
    if "내부보안감사체크리스트" in name:
        sheet_name = name
        break
if not sheet_name:
    sheet_name = wb.sheetnames[3]

sheet = wb[sheet_name]
print(f"Sheet Title: {sheet.title}")

# 결함 판단 로직
def is_red_font(cell):
    if not cell or not cell.font or not cell.font.color:
        return False
    color = cell.font.color
    if color.type == 'rgb' and color.rgb:
        rgb_val = str(color.rgb).upper()
        if len(rgb_val) == 8:
            rgb_val = rgb_val[2:]
        if len(rgb_val) == 6:
            try:
                r = int(rgb_val[0:2], 16)
                g = int(rgb_val[2:4], 16)
                b = int(rgb_val[4:6], 16)
                if r > 180 and g < 110 and b < 110:
                    return True
            except ValueError:
                pass
    if color.indexed in (10, 2):
        return True
    return False

rows_info = []
for r in range(3, sheet.max_row + 1):
    item_name = sheet.cell(row=r, column=3).value
    detail = sheet.cell(row=r, column=4).value
    eval_val = sheet.cell(row=r, column=5).value
    status = sheet.cell(row=r, column=6).value
    improv = sheet.cell(row=r, column=7).value
    
    status_cell = sheet.cell(row=r, column=6)
    improv_cell = sheet.cell(row=r, column=7)
    
    is_status_red = is_red_font(status_cell)
    is_improv_red = is_red_font(improv_cell)
    
    rows_info.append({
        "row_idx": r,
        "item_name": item_name,
        "eval": eval_val,
        "is_status_red": is_status_red,
        "is_improv_red": is_improv_red,
        "improv": improv,
        "status": status
    })

print("Total Rows scanned:", len(rows_info))
red_status_rows = [r for r in rows_info if r["is_status_red"]]
red_improv_rows = [r for r in rows_info if r["is_improv_red"]]
x_eval_rows = [r for r in rows_info if r["eval"] == "X"]

print("Status Red font rows:", [r["row_idx"] for r in red_status_rows])
print("Improv Red font rows:", [r["row_idx"] for r in red_improv_rows])
print("Eval 'X' rows:", [r["row_idx"] for r in x_eval_rows])

# "X" 평가를 가진 행의 상세 정보 출력
print("\n--- Eval 'X' Rows Details ---")
for r in x_eval_rows:
    print(f"Row {r['row_idx']} | Item: {r['item_name']} | Eval: {r['eval']}")
    print(f"  Status: {r['status']}")
    print(f"  Improv: {r['improv']}")

print("\n--- Improv Red Rows Details ---")
for r in red_improv_rows:
    print(f"Row {r['row_idx']} | Item: {r['item_name']} | Eval: {r['eval']}")
    print(f"  Status: {r['status']}")
    print(f"  Improv: {r['improv']}")
