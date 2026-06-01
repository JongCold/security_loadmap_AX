import os
import openpyxl
import sys

# 인코딩 설정
sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
result_file = os.path.join(BASE_DIR, "2026년 KG그룹_KG이니시스 정보보안감사_보안솔루션로드맵_20260601.xlsx")

if not os.path.exists(result_file):
    print(f"Error: File not found {result_file}")
    exit(1)

wb = openpyxl.load_workbook(result_file, data_only=True)
sheet_name = None
for name in wb.sheetnames:
    if "01_KG그룹" in name or "01_" in name:
        sheet_name = name
        break
if not sheet_name:
    sheet_name = wb.sheetnames[1]

sheet = wb[sheet_name]
print(f"Result Sheet Name: {sheet.title}")
print(f"Max row: {sheet.max_row}")

print("\nMerged cells in result file (row >= 20):")
merges = sorted(list(sheet.merged_cells.ranges), key=lambda r: r.min_row)
for merge in merges:
    if merge.min_row >= 20:
        print(merge)

print("\nRoadmap Data Rows (from row 140 to max_row):")
# 자산목록이 추가되어서 로드맵 데이터가 140행 이후로 밀렸을 수 있습니다.
# 헤더 이후의 로드맵 시작 위치를 찾아봅니다.
# 로드맵 헤더는 보통 "No." 이고 2열은 "항목명" 입니다.
# 24행 근처에 "No." 가 있었는데 자산목록이 100행 이상 추가되었으므로 헤더도 밀렸을 것입니다.
header_row = None
for r in range(1, sheet.max_row + 1):
    val = sheet.cell(row=r, column=1).value
    if val == "No." or val == "No":
        # 11행의 No (자산목록 헤더) 외에 다음 No. 가 로드맵 헤더입니다.
        if r > 15:
            header_row = r
            print(f"Found Roadmap Header at Row {r}")

if header_row:
    for r in range(header_row, min(header_row + 30, sheet.max_row + 1)):
        val_a = sheet.cell(row=r, column=1).value
        val_b = sheet.cell(row=r, column=2).value
        val_c = sheet.cell(row=r, column=3).value
        val_d = sheet.cell(row=r, column=4).value
        val_e = sheet.cell(row=r, column=5).value
        val_f = sheet.cell(row=r, column=6).value
        val_g = sheet.cell(row=r, column=7).value
        val_h = sheet.cell(row=r, column=8).value
        val_i = sheet.cell(row=r, column=9).value
        val_j = sheet.cell(row=r, column=10).value
        val_p = sheet.cell(row=r, column=16).value
        
        print(f"Row {r:03d} | Col1: {val_a} | Col2: {str(val_b)[:15]} | Col3-4: C={str(val_c)[:10]}, D={str(val_d)[:10]} | Col5-6: E={str(val_e)[:10]}, F={str(val_f)[:10]} | Col7-8: G={str(val_g)[:10]}, H={str(val_h)[:10]} | Col9: {val_i} | Col10: {str(val_j)[:10]} | Col16 (Note): {repr(val_p)[:40]}")
else:
    print("Roadmap header not found by keyword 'No.' or 'No'. Listing raw rows from 140 to 180:")
    for r in range(140, min(180, sheet.max_row + 1)):
        val_a = sheet.cell(row=r, column=1).value
        val_b = sheet.cell(row=r, column=2).value
        print(f"Row {r:03d} | Col1: {val_a} | Col2: {val_b}")
