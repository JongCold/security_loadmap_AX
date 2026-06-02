import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_formatted_output.xlsx")

wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active
print(f"Sheet Name: {sheet.title}")

# 로드맵 데이터가 시작되는 영역 찾기 (보안솔루션로드맵 헤더 아래)
# 26행 근처가 데이터 시작점일 것
print("\n--- Rows Information ---")
for r in range(23, sheet.max_row + 1):
    val_a = sheet.cell(row=r, column=1).value
    val_b = sheet.cell(row=r, column=2).value
    val_c = sheet.cell(row=r, column=3).value
    val_g = sheet.cell(row=r, column=7).value
    val_p = sheet.cell(row=r, column=16).value
    
    height = sheet.row_dimensions[r].height
    
    if val_a or val_b or val_c:
        print(f"Row {r} (Height: {height}):")
        print(f"  Col A (No): {val_a}")
        print(f"  Col B (항목명): {val_b}")
        print(f"  Col C (세부점검내용): {str(val_c)[:40]}...")
        
        # 7열 개선방안 폰트 정보
        cell_g = sheet.cell(row=r, column=7)
        g_color = cell_g.font.color.rgb if cell_g.font and cell_g.font.color else "No Color"
        print(f"  Col G (개선방안): {str(val_g)[:40]}... (Color RGB: {g_color})")
        
        # 16열 비고 폰트 정보
        cell_p = sheet.cell(row=r, column=16)
        p_color = cell_p.font.color.rgb if cell_p.font and cell_p.font.color else "No Color"
        print(f"  Col P (비고):\n{val_p}\n  (Color RGB: {p_color})")
        print("-" * 50)

print("\n--- Column Widths ---")
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
    print(f"Col {col} Width: {sheet.column_dimensions[col].width}")
