import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
asset_file = os.path.join(BASE_DIR, "KG그룹(KG가족사명)자산목록 및 중요도 평가서_양식.xlsx")

wb = openpyxl.load_workbook(asset_file, data_only=True)
print("Sheet Names:", wb.sheetnames)

# '4.정보보호시스템' 시트 혹은 이름에 '정보보호시스템'이 들어가는 시트 선택
sheet = None
for name in wb.sheetnames:
    if "정보보호시스템" in name.replace(" ", ""):
        sheet = wb[name]
        break
if not sheet:
    sheet = wb.active

print(f"Active Sheet: {sheet.title}")

# 헤더 스캔을 위해 7, 8행의 컬럼 값 출력
for r in [7, 8]:
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 20)]
    print(f"Row {r}: {row_vals}")

# 예시 데이터 9행 출력
print(f"Row 9: {[sheet.cell(row=9, column=c).value for c in range(1, 20)]}")
