import sys

# UTF-8 출력 재설정
sys.stdout.reconfigure(encoding='utf-8')

# 정상 상태의 1~279행 코드 조각과 280~371행 코드 조각, 그리고 고도화된 generate_roadmap_excel 정의를 합쳐서 
# roadmap_agent_llm_new.py를 새로 쓰는 스크립트.

part1_up_to_279 = """import os
import re
import sys
import json
import collections
import math
from copy import copy
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import requests

# RAG 엔진 임포트
from rag_engine import get_rag_engine, initialize_rag

# 표준 출력 인코딩을 UTF-8로 설정 및 라인 버퍼링 강제
sys.stdout.reconfigure(encoding='utf-8', line_buffering=True)

# 경로 설정 - 프로젝트 디렉토리 기준 상대 경로 자동화
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SOLUTION_LIST_PATH = os.path.join(BASE_DIR, "자사 보안 솔루션 리스트 (1).xlsx")
ROADMAP_TEMPLATE_PATH = os.path.join(BASE_DIR, "2026년 KG그룹_제로인 정보보안감사_보안솔루션로드맵_v1.2.xlsx")

# 로컬 Ollama 설정
OLLAMA_URL = "http://127.0.0.1:11434/api/generate"
OLLAMA_MODEL = "gemma2:2b"

def is_red_font(cell):
    \"\"\"셀의 글꼴 색상이 빨간색 계열인지 여부를 판별\"\"\"
    if not cell or not cell.font or not cell.font.color:
        return False
    color = cell.font.color
    
    # 1. RGB 값 직접 검사
    if color.type == 'rgb' and color.rgb:
        rgb_val = str(color.rgb).upper()
        if len(rgb_val) == 8:
            rgb_val = rgb_val[2:]  # Alpha 채널 제거
        if len(rgb_val) == 6:
            try:
                r = int(rgb_val[0:2], 16)
                g = int(rgb_val[2:4], 16)
                b = int(rgb_val[4:6], 16)
                # 빨간색 채널이 압도적으로 높고 녹색/청색이 낮은 조건
                if r > 180 and g < 110 and b < 110:
                    return True
            except ValueError:
                pass
                
    # 2. 테마 또는 다른 타입인 경우 빨간색 인덱스 검사 (보완용)
    if color.indexed == 10 or color.indexed == 2:  # Excel 표준 Red 인덱스
        return True
        
    return False


def is_red_fill(cell):
    \"\"\"셀의 배경색이 빨간색 계열인지 여부를 판별\"\"\"
    if not cell or not cell.fill or not cell.fill.start_color:
        return False
    color = cell.fill.start_color
    
    # 1. RGB 값 직접 검사
    if color.type == 'rgb' and color.rgb:
        rgb_val = str(color.rgb).upper()
        if len(rgb_val) == 8:
            rgb_val = rgb_val[2:]  # Alpha 채널 제거
        if len(rgb_val) == 6:
            try:
                r = int(rgb_val[0:2], 16)
                g = int(rgb_val[2:4], 16)
                b = int(rgb_val[4:6], 16)
                # 빨간색 채널이 압도적으로 높고 녹색/청색이 낮은 조건
                if r > 180 and g < 110 and b < 110:
                    return True
            except ValueError:
                pass
                
    # 2. 테마 또는 다른 타입인 경우 빨간색 인덱스 검사 (보완용)
    if color.indexed == 10 or color.indexed == 2:  # Excel 표준 Red 인덱스
        return True
        
    return False


def find_best_matching_solution(item):
    \"\"\"ChromaDB RAG 기반 자사 솔루션 최적 매핑

    기존 단순 코사인 유사도 \u2192 ChromaDB 벡터 임베딩 기반 검색으로 대체.
    자사 보안 솔루션 리스트의 벡터 DB에서 결함 항목과 가장 유사한 솔루션을 검색합니다.
    \"\"\"
    try:
        rag = get_rag_engine()
        best_sol, similarity = rag.find_best_solution(item, similarity_threshold=0.20)

        if best_sol:
            return {
                "보안영역": best_sol.get("보안영역", ""),
                "솔루션구분": best_sol.get("솔루션구분", ""),
                "제조사명": best_sol.get("제조사명", ""),
                "제품명": best_sol.get("제품명", ""),
                "제품명기능설명": best_sol.get("제품명기능설명", "")
            }, similarity
        else:
            return None, similarity
    except Exception as e:
        print(f"[RAG] 매핑 에러: {e}", flush=True)
        return None, 0.0


def parse_red_cells_from_checklist(filepath):
    \"\"\"상세체크리스트 파일에서 조치가 필요한 결함 사항(평가 'X') 및 기존에 입력된 데이터 추출\"\"\"
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"체크리스트 파일을 찾을 수 없습니다: {filepath}")
        
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if "내부보안감사체크리스트" in name:
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[3]
        
    sheet = wb[sheet_name]
    defect_items = []
    
    last_item_name = "미정" # 병합 셀(Col 3 항목명) 처리를 위한 직전 유효 항목명 저장 변수
    
    for r in range(3, sheet.max_row + 1):
        item_name_cell = sheet.cell(row=r, column=3) # 항목명 (Col 3)
        detail_cell = sheet.cell(row=r, column=4) # 세부점검내용 (Col 4)
        eval_cell = sheet.cell(row=r, column=5) # 평가 (Col 5)
        status_cell = sheet.cell(row=r, column=6) # 운영현황 및 증적 (Col 6)
        improv_cell = sheet.cell(row=r, column=7) # 개선방안 (Col 7)
        
        eval_val = str(eval_cell.value).strip().upper() if eval_cell.value else ""
        status_val = str(status_cell.value).strip() if status_cell.value else ""
        improv_val = str(improv_cell.value).strip() if improv_cell.value else ""
        
        # C열(항목명) 셀이 병합되어 비어있는(None) 경우 직전 행의 항목명을 승계하여 바인딩
        if item_name_cell.value is not None and str(item_name_cell.value).strip() != "":
            last_item_name = str(item_name_cell.value).strip()
            
        # 1차 결함 판정: 평가 열(E열, Col 5)의 값이 정확히 "X"인 경우 조치 대상 결함으로 판단
        is_defect = (eval_val == "X")
        
        if is_defect:
            # 기존에 채워져 있을 수 있는 나머지 열의 정보들 파싱 (UI 표출용)
            area_val = str(sheet.cell(row=r, column=9).value).strip() if sheet.cell(row=r, column=9).value else "기타 보안"
            project_val = str(sheet.cell(row=r, column=10).value).strip() if sheet.cell(row=r, column=10).value else "보안 솔루션 구축"
            law_val = str(sheet.cell(row=r, column=11).value).strip() if sheet.cell(row=r, column=11).value else "N/A"
            urgency_val = safe_int(sheet.cell(row=r, column=12).value, 3)
            risk_val = safe_int(sheet.cell(row=r, column=13).value, 3)
            budget_val = str(sheet.cell(row=r, column=14).value).strip() if sheet.cell(row=r, column=14).value else "영업 문의가 필요한 영역 논의 필요."
            year_val = str(sheet.cell(row=r, column=15).value).strip() if sheet.cell(row=r, column=15).value else "2026년"
            note_val = str(sheet.cell(row=r, column=16).value).strip() if sheet.cell(row=r, column=16).value else ""

            # 비고란에서 제조사 및 추천솔루션 파싱
            rec_sol = "N/A"
            rec_vendor = "N/A"
            if note_val:
                # [추천 솔루션] 제조사 - 제품명
                match = re.search(r'\[추천\s*솔루션\]\s*(.*?)\s*-\s*([^\n\r]+)', note_val)
                if match:
                    rec_vendor = match.group(1).strip()
                    rec_sol = match.group(2).strip()
                elif "N/A" in note_val:
                    rec_sol = "N/A"
                    rec_vendor = "N/A"

            defect_items.append({
                "row_idx": r,
                "항목명": last_item_name,
                "세부점검내용": str(detail_cell.value).strip() if detail_cell.value else "",
                "평가": eval_val,
                "운영현황_증적": status_val,
                "개선방안": improv_val,
                "보안영역": area_val,
                "과제명": project_val,
                "법적요구": law_val,
                "시급성": urgency_val,
                "위험도": risk_val,
                "예상예산": budget_val,
                "로드맵연도": year_val,
                "비고": note_val,
                "추천솔루션": rec_sol,
                "제조사": rec_vendor
            })
            
    return defect_items


def safe_int(val, default=3):
    \"\"\"안전한 정수 형변환 헬퍼 함수 (정수 변환 실패 시 기본값 반환)\"\"\"
    try:
        if val is None:
            return default
        if isinstance(val, (int, float)):
            return int(val)
        nums = re.findall(r'\d+', str(val))
        if nums:
            return int(nums[0])
        return default
    except Exception:
        return default


def call_local_gemma(item, sol_context=None, model_name=OLLAMA_MODEL):
    \"\"\"로컬 Ollama LLM 모델을 사용하여 솔루션 자동 추천 및 매핑 수행 (최대 3회 재시도)
    
    [하이브리드 RAG + LLM 오케스트레이션]
    1단계: ChromaDB RAG에서 결함 항목에 최적의 자사 솔루션 검색
    2단계: RAG 검색 결과를 Context로 LLM 프롬프트에 주입
    3단계: LLM이 Context + Query를 조합하여 최종 JSON 결과 생성
    \"\"\"
    # ============================================================
    # 1단계: RAG 기반 자사 솔루션 최적 매핑 탐색 (ChromaDB 벡터 검색)
    # ============================================================
    best_sol, score = find_best_matching_solution(item)
    print(f"[RAG] 매핑 결과: 제품={best_sol['제품명'] if best_sol else 'N/A'}, "
          f"제조사={best_sol['제조사명'] if best_sol else 'N/A'}, "
          f"유사도={score:.4f}", flush=True)
    
    # ============================================================
    # 2단계: RAG 컨텍스트(Markdown) + Query(JSON) 하이브리드 프롬프트 구성
    # ============================================================
    
    # RAG 검색으로 Top-3 솔루션의 Markdown 컨텍스트도 가져옴
    rag = get_rag_engine()
    rag_context_md = rag.get_rag_context_markdown(item, top_k=3)
    
    if best_sol and score >= 0.20:
        rag_guideline = f\"\"\"
[RAG 매핑된 실제 자사 보안 솔루션 지정]
너는 반드시 아래에 명시된 자사 보안 솔루션의 정보만을 활용하여 결과를 작성해야 한다.
- 보안영역: {best_sol['보안영역']}
- 솔루션 구분: {best_sol['솔루션구분']}
- 제조사명: {best_sol['제조사명']}
- 제품명: {best_sol['제품명']}
- 제품 기능 설명: {best_sol['제품명기능설명']}

[지시사항]
- "보안영역" 키 값에는 반드시 위 솔루션의 실제 보안영역("{best_sol['보안영역']}")을 기재할 것.
- "과제명" 키 값에는 구체적으로 위 솔루션의 제품명("{best_sol['제품명']}") 도입 목적을 문장으로 완성해 작성할 것. (예: "{best_sol['제품명']} 도입 및 보안 통제 강화")
- "비고" 키 값에는 아래 서식을 정확하게 준수하여 기재할 것.
  [추천 솔루션] {best_sol['제조사명']} - {best_sol['제품명']}
  [선정 이유] {best_sol['제품명기능설명']}의 기능을 바탕으로 고객사의 결함 사항(운영현황 및 개선방안)을 어떻게 구체적/기술적으로 조치할 수 있는지 설명.

{rag_context_md}
\"\"\"
    else:
        rag_guideline = f\"\"\"
[RAG 매핑 결과: 적절한 도입 보안 솔루션 없음]
이 결함은 특정 하드웨어나 소프트웨어 보안 장비 도입이 불필요한 단순 정책 수립, 프로세스 마련, 또는 수동 감사용 성격의 항목입니다.
따라서 너는 반드시 아래의 규칙에 따라 결과 JSON을 작성해야 한다.
- "보안영역" 키 값에는 "기타 보안" 또는 적절한 일반 관리 보안 영역을 기재할 것.
- "과제명" 키 값에는 "보안 통제 절차 수립 및 규정 강화"라고 정확히 기재할 것.
- "비고" 키 값에는 아래 서식을 정확하게 준수하여 기재할 것.
  [추천 솔루션] N/A
  [선정 이유] 자사 보안 솔루션 풀 내에 해당하는 기술적 솔루션이 존재하지 않으므로, 장비 도입 대신 사내 보안 지침/절차 수립 및 수동 관리 감독 프로세스 강화를 권고합니다.

{rag_context_md}
\"\"\"

    # ============================================================
    # 3단계: 프롬프트 오케스트레이션 (CoT: Chain of Thought 적용)
    # ============================================================
    prompt = f\"\"\"
역할: 정보보안 CISO 컨설턴트 및 솔루션 아키텍트.
주어진 [고객사 결함사항]을 정밀 분석하고, [자사 보안 RAG 매핑 가이드라인]을 근거로 하여 지정된 JSON 구조로 최종 결과를 수립하라.

{rag_guideline}

[고객사 결함사항]
- 진단 항목명: {item['항목명']}
- 점검 세부내용: {item['세부점검내용']}
- 실태 운영현황: {item['운영현황_증적']}
- 권고 개선방안: {item['개선방안']}

[추가 제약 조건]
1. 법적요구: 고객사의 결함사항(진단 항목명, 점검 세부내용, 실태 운영현황, 권고 개선방안) 텍스트 내부에서 "개인정보보호법", "정보통신망법", "망법" 또는 구체적 법적 조항(예: 제29조 등)이 직접 언급된 경우에만 해당 법적 조항명을 적으십시오. 만약 직접적인 법령 언급이 없다면 무조건 "N/A"로 설정해야 합니다. (지어내거나 유추하지 마십시오.)
2. 예상예산: 도입 제품이 존재할 시 도입 규모를 고려해 현실적인 예산 범위(예: ₩30,000,000)를 기입하고, 도입 제품이 N/A인 경우는 무조건 "영업 문의가 필요한 영역 논의 필요." 라고 적으십시오.
3. 임의의 플레이스홀더 단어(예: "[실제 제품명]", "[사용자 정의]", "[사업/개발]")를 최종 텍스트에 노출하는 행위는 절대 금지합니다.

[출력 스펙 요구사항]
오직 아래의 Key 값을 포함한 단 하나의 JSON 데이터만 반환하라. 어떠한 앞뒤 설명이나 마크다운 백틱(```json)도 포함하면 안 된다.
{{
  "보안영역": "RAG 가이드라인에 지정된 보안영역",
  "과제명": "RAG 가이드라인에 지정된 과제명",
  "법적요구": "언급된 법적 요건 명시 또는 무조건 'N/A'",
  "시급성": 1에서 5 사이의 정수 (법적 의무 사항이 명시된 경우 5로 설정)",
  "위험도": 1에서 5 사이의 정수",
  "예상예산": "₩30,000,000 등 도입 비용 또는 '영업 문의가 필요한 영역 논의 필요.'",
  "비고": "RAG 가이드라인에 지정된 서식에 따라 작성된 비고 내용"
}}
\"\"\"
    max_retries = 3
    for attempt in range(1, max_retries + 1):
        try:
            payload = {
                "model": model_name,
                "prompt": prompt,
                "stream": False,
                "format": "json" # JSON 모드 강제
            }
            # 타임아웃을 180초로 증가하여 CPU 환경 등에서의 타임아웃 방지
            res = requests.post(OLLAMA_URL, json=payload, timeout=180)
            if res.status_code == 200:
                raw_text = res.json().get("response", "").strip()
                # 줄바꿈 이스케이프 오류 등 방지용 정규화
                clean_text = re.sub(r'[\r\n]+', ' ', raw_text)
                data = json.loads(clean_text)
                
                # RAG 조회 결과로 얻은 실제 제조사명과 제품명을 추가로 탑재
                if best_sol and score >= 0.20:
                    data["추천솔루션"] = best_sol.get("제품명", "N/A")
                    data["제조사"] = best_sol.get("제조사명", "N/A")
                else:
                    data["추천솔루션"] = "N/A"
                    data["제조사"] = "N/A"
                return data
            else:
                raise Exception(f"Ollama API 리턴 에러: {res.status_code}")
        except Exception as e:
            print(f"[시도 {attempt}/{max_retries}] Ollama 호출 실패: {e}", flush=True)
            if attempt == max_retries:
                print("최대 재시도 횟수를 초과했습니다. 수동 검토용 템플릿으로 대체합니다.", flush=True)
                
    # 최종 실패 시 반환할 수동 검토용 기본 데이터
    return {
        "보안영역": "미지정",
        "과제명": "솔루션 도입 검토 필요 (분석 실패)",
        "법적요구": "N/A",
        "시급성": 3,
        "위험도": 3,
        "예상예산": "영업 문의가 필요한 영역 논의 필요.",
        "비고": f"[추천 솔루션] 분석 실패\\n[선정 이유] LLM 서버 미기동 또는 응답 형식 분석 오류로 분석을 완료하지 못했습니다. 수동 검토가 필요합니다.",
        "추천솔루션": "N/A",
        "제조사": "N/A"
    }

def calculate_roadmap_year(res):
    \"\"\"시급성, 위험도, 법적요구 기준 연도 분배\"\"\"
    urgency = safe_int(res.get("시급성", 3))
    risk = safe_int(res.get("위험도", 3))
    law = res.get("법적요구", "N/A")
    
    if law and law != "N/A":
        return "2026년"
        
    score = urgency * risk
    if score >= 16:
        return "2026년"
    elif score >= 8:
        return "2027년"
    else:
        return "2028년"

def copy_cell_style(src_cell, dest_cell):
    \"\"\"셀 스타일 서식 복사\"\"\"
    if src_cell.font:
        dest_cell.font = copy(src_cell.font)
    if src_cell.border:
        dest_cell.border = copy(src_cell.border)
    if src_cell.fill:
        dest_cell.fill = copy(src_cell.fill)
    if src_cell.alignment:
        dest_cell.alignment = copy(src_cell.alignment)
    if src_cell.number_format:
        dest_cell.number_format = copy(src_cell.number_format)
"""

part2_generate_roadmap_excel = """
def generate_roadmap_excel(results, company_name, output_filepath, env_filepath=None, asset_filepath=None):
    \"\"\"원본 로드맵 템플릿 서식을 유지하며 매핑 결과 및 사전환경조사서, 자산목록 데이터를 통합하여 엑셀 작성 및 저장\"\"\"
    if not os.path.exists(ROADMAP_TEMPLATE_PATH):
        raise FileNotFoundError(f"템플릿 엑셀 파일을 찾을 수 없습니다: {ROADMAP_TEMPLATE_PATH}")
        
    wb = openpyxl.load_workbook(ROADMAP_TEMPLATE_PATH)
    sheet_name = None
    for name in wb.sheetnames:
        if "01_KG그룹" in name or "01_" in name:
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[1]
        
    sheet = wb[sheet_name]
    
    # 조작 시작 전 원래 시트의 행 높이 정보를 딕셔너리로 백업
    original_heights = {}
    for r in range(1, sheet.max_row + 1):
        if r in sheet.row_dimensions:
            original_heights[r] = sheet.row_dimensions[r].height
            
    # 0. 병합 셀 백업 및 시트 전체 병합 해제
    original_merges = list(sheet.merged_cells.ranges)
    for m_range in original_merges:
        sheet.unmerge_cells(start_row=m_range.min_row, start_column=m_range.min_col, 
                            end_row=m_range.max_row, end_column=m_range.max_col)
    
    # 1. 사전환경조사서 병합 반영
    if env_filepath and os.path.exists(env_filepath):
        print(f"[엑셀 병합] 사전환경조사서 반영 시작: {env_filepath}", flush=True)
        try:
            env_wb = openpyxl.load_workbook(env_filepath, data_only=True)
            env_sheet = None
            for name in env_wb.sheetnames:
                if "KG그룹" in name:
                    env_sheet = env_wb[name]
                    break
            if not env_sheet:
                env_sheet = env_wb.worksheets[2] if len(env_wb.worksheets) > 2 else env_wb.active
                
            def clean_env_value(val, unit=""):
                if val is None:
                    return ""
                val_str = str(val).strip()
                if val_str == "" or val_str.lower() == "none":
                    return ""
                if unit and not val_str.endswith(unit):
                    try:
                        num = float(val_str)
                        if num.is_integer():
                            return f"{int(num):,}{unit}"
                        return f"{num:,}{unit}"
                    except ValueError:
                        return f"{val_str}{unit}"
                return val_str
                
            # 값 읽기
            c_name = clean_env_value(env_sheet.cell(row=5, column=2).value)
            b_field = clean_env_value(env_sheet.cell(row=5, column=8).value)
            b_type = clean_env_value(env_sheet.cell(row=6, column=2).value)
            is_tele = clean_env_value(env_sheet.cell(row=6, column=8).value)
            emp_cnt = clean_env_value(env_sheet.cell(row=7, column=2).value, "명")
            revenue = clean_env_value(env_sheet.cell(row=7, column=8).value)
            laws = clean_env_value(env_sheet.cell(row=8, column=2).value)
            gov = clean_env_value(env_sheet.cell(row=8, column=8).value)
            
            # 템플릿에 쓰기
            if c_name:
                sheet.cell(row=5, column=2, value=c_name)
            if b_field:
                sheet.cell(row=5, column=9, value=b_field)
            if b_type:
                sheet.cell(row=6, column=2, value=b_type)
            if is_tele:
                sheet.cell(row=6, column=9, value=is_tele)
            if emp_cnt:
                sheet.cell(row=7, column=2, value=emp_cnt)
            if revenue:
                sheet.cell(row=7, column=9, value=revenue)
            if laws:
                sheet.cell(row=8, column=2, value=laws)
            if gov:
                sheet.cell(row=8, column=9, value=gov)
                
            print("[엑셀 병합] 사전환경조사서 상단 영역 기입 완료", flush=True)
        except Exception as env_err:
            print(f"[엑셀 병합] 사전환경조사서 파싱 중 에러 발생: {env_err}", flush=True)

    # 2. 자산목록 데이터 파싱 및 보안 솔루션 세부 현황 동적 행 삽입/삭제
    diff = 0
    N = 0
    if asset_filepath and os.path.exists(asset_filepath):
        print(f"[엑셀 병합] 자산목록 반영 시작: {asset_filepath}", flush=True)
        try:
            asset_wb = openpyxl.load_workbook(asset_filepath, data_only=True)
            asset_sheet = None
            for name in asset_wb.sheetnames:
                if "4.정보보호시스템" in name:
                    asset_sheet = asset_wb[name]
                    break
            if not asset_sheet:
                for sh in asset_wb.worksheets:
                    if "정보보호시스템" in sh.title:
                        asset_sheet = sh
                        break
            if not asset_sheet:
                asset_sheet = asset_wb.worksheets[10] if len(asset_wb.worksheets) > 10 else asset_wb.active
                
            asset_rows = []
            for r in range(9, asset_sheet.max_row + 1):
                no_val = asset_sheet.cell(row=r, column=2).value
                if no_val is None:
                    continue
                try:
                    no_int = int(no_val)
                except ValueError:
                    continue
                    
                gubun = asset_sheet.cell(row=r, column=4).value
                ip = asset_sheet.cell(row=r, column=5).value
                host = asset_sheet.cell(row=r, column=6).value
                vendor = asset_sheet.cell(row=r, column=7).value
                func = asset_sheet.cell(row=r, column=8).value
                location = asset_sheet.cell(row=r, column=9).value
                manager = asset_sheet.cell(row=r, column=10).value
                dept = asset_sheet.cell(row=r, column=12).value
                
                c_val = asset_sheet.cell(row=r, column=13).value
                i_val = asset_sheet.cell(row=r, column=14).value
                a_val = asset_sheet.cell(row=r, column=15).value
                
                def clean(v):
                    return str(v).strip() if v is not None else ""
                
                def to_int(v, default=3):
                    if v is None:
                        return default
                    try:
                        return int(v)
                    except ValueError:
                        nums = re.findall(r'\d+', str(v))
                        if nums:
                            return int(nums[0])
                        return default

                c_num = to_int(c_val, 3)
                i_num = to_int(i_val, 3)
                a_num = to_int(a_val, 3)
                tot_num = c_num + i_num + a_num
                
                # 등급 산정: 합산 7~9 = H, 5~6 = M, 3~4 = L
                if tot_num >= 7:
                    grade = "H"
                elif tot_num >= 5:
                    grade = "M"
                else:
                    grade = "L"
                
                gubun_str = clean(gubun)
                ip_str = clean(ip)
                host_str = clean(host)
                vendor_str = clean(vendor)
                
                # 구분, IP 주소, 호스트명, 벤더 및 모델명이 모두 비어있는 행은 불필요하므로 제거(스킵)
                if not gubun_str and not ip_str and not host_str and not vendor_str:
                    continue
                    
                # 빨간색 바탕 셀이 있는 자산인지 검사
                is_inactive = False
                for col_idx in range(1, 18):
                    cell_to_check = asset_sheet.cell(row=r, column=col_idx)
                    if is_red_fill(cell_to_check):
                        is_inactive = True
                        break
                
                asset_rows.append({
                    "구분": gubun_str,
                    "IP주소": ip_str,
                    "벤더": vendor_str,
                    "기능": clean(func),
                    "자산위치": clean(location),
                    "담당자": clean(manager),
                    "담당부서": clean(dept),
                    "기밀성": c_num,
                    "무결성": i_num,
                    "가용성": a_num,
                    "등급": grade,
                    "비고": "운영하고 있지 않는 자산" if is_inactive else "",
                    "fills": {
                        2: asset_sheet.cell(row=r, column=4).fill,  # 구분
                        3: asset_sheet.cell(row=r, column=5).fill,  # IP주소
                        4: asset_sheet.cell(row=r, column=7).fill,  # 벤더
                        5: asset_sheet.cell(row=r, column=8).fill,  # 기능
                        6: asset_sheet.cell(row=r, column=9).fill,  # 자산위치
                        7: asset_sheet.cell(row=r, column=10).fill, # 담당자
                        8: asset_sheet.cell(row=r, column=12).fill, # 담당부서
                        9: asset_sheet.cell(row=r, column=13).fill,  # 기밀성
                        10: asset_sheet.cell(row=r, column=14).fill, # 무결성
                        11: asset_sheet.cell(row=r, column=15).fill, # 가용성
                        12: asset_sheet.cell(row=r, column=16).fill, # 합계
                        13: asset_sheet.cell(row=r, column=17).fill, # 등급
                        14: asset_sheet.cell(row=r, column=17).fill, # 비고
                    }
                })
                
            N = len(asset_rows)
            print(f"[엑셀 병합] 자산목록에서 필터링 후 유효 보안장비 {N}건 검출", flush=True)
            
            if N > 0:
                diff = N - 10
                
                # 행 조정 수행 (병합이 해제되어 있으므로 자유롭게 행 삽입/삭제 가능)
                if diff > 0:
                    sheet.insert_rows(23, amount=diff)
                    for r in range(23, 23 + diff):
                        for col_idx in range(1, 17):
                            src_cell = sheet.cell(row=13, column=col_idx)
                            dest_cell = sheet.cell(row=r, column=col_idx)
                            copy_cell_style(src_cell, dest_cell)
                elif diff < 0:
                    sheet.delete_rows(13 + N, amount=-diff)
                    
                # 행 높이 일정 정렬 및 오프셋 보정 복구
                default_height = original_heights.get(13, 22)
                for r in range(13, 13 + N):
                    sheet.row_dimensions[r].height = default_height
 
                # 자산이 들어간 각 셀의 너비를 가독성 좋게 일정 크기 조정
                col_widths = {
                    1: 6,    # No.
                    2: 18,   # 구분
                    3: 20,   # IP주소
                    4: 25,   # 벤더 및 모델명
                    5: 30,   # 기능
                    6: 18,   # 자산위치
                    7: 12,   # 담당자
                    8: 15,   # 담당부서
                    9: 10,   # 기밀성
                    10: 10,  # 무결성
                    11: 10,  # 가용성
                    12: 10,  # 합계
                    13: 10,  # 등급
                }
                for col_idx, width in col_widths.items():
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    sheet.column_dimensions[col_letter].width = width
                    
                # 밀려난 23행 이상의 행들에 백업된 행 높이를 오프셋 반영하여 재지정
                for r, h in original_heights.items():
                    if r < 13:
                        sheet.row_dimensions[r].height = h
                    elif r >= 23:
                        new_r = r + diff
                        sheet.row_dimensions[new_r].height = h
                    
                # 기밀성, 무결성, 가용성, 합계, 등급, 비고 헤더로 셀 분류 기입 및 스타일 일관성 복사
                header_style_cell = sheet.cell(row=12, column=2)
                center_align = Alignment(horizontal="center", vertical="center")
                
                for col_idx in range(9, 17):
                    copy_cell_style(header_style_cell, sheet.cell(row=12, column=col_idx))
                    sheet.cell(row=12, column=col_idx).alignment = center_align
 
                sheet.cell(row=12, column=9, value="기밀성")
                sheet.cell(row=12, column=10, value="무결성")
                sheet.cell(row=12, column=11, value="가용성")
                sheet.cell(row=12, column=12, value="합계")
                sheet.cell(row=12, column=13, value="등급")
                sheet.cell(row=12, column=14, value="비고")
                
                # 신규 세부 현황 행들에 대한 개별 값 기입
                for i, asset in enumerate(asset_rows):
                    r = 13 + i
                    std_border = sheet.cell(row=r, column=8).border
                    data_center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    row_fill = asset["fills"].get(2)
                    
                    cols_map = {
                        1: (i + 1, None),
                        2: (asset["구분"], asset["fills"].get(2)),
                        3: (asset["IP주소"], asset["fills"].get(3)),
                        4: (asset["벤더"], asset["fills"].get(4)),
                        5: (asset["기능"], asset["fills"].get(5)),
                        6: (asset["자산위치"], asset["fills"].get(6)),
                        7: (asset["담당자"], asset["fills"].get(7)),
                        8: (asset["담당부서"], asset["fills"].get(8)),
                        9: (asset["기밀성"], asset["fills"].get(9)),
                        10: (asset["무결성"], asset["fills"].get(10)),
                        11: (asset["가용성"], asset["fills"].get(11)),
                        12: (f"=SUM(I{r}:K{r})", asset["fills"].get(12)),
                        13: (asset["등급"], asset["fills"].get(13)),
                        14: (asset["비고"], asset["fills"].get(14))
                    }
                    
                    for col_idx, (val, src_fill) in cols_map.items():
                        cell = sheet.cell(row=r, column=col_idx, value=val)
                        if col_idx >= 9:
                            cell.alignment = data_center_align
                            if std_border:
                                cell.border = copy(std_border)
                                
                        if src_fill and src_fill.fill_type and src_fill.fill_type != 'none':
                            cell.fill = copy(src_fill)
                        elif row_fill and row_fill.fill_type and row_fill.fill_type != 'none' and col_idx >= 9:
                            cell.fill = copy(row_fill)
                    
                print(f"[엑셀 병합] 자산목록 {N}건 데이터 바인딩 및 배경색 이식 완료 (행 조정 오프셋: {diff})", flush=True)
            else:
                print("[엑셀 병합] 자산목록에 유효한 데이터가 없어 세부 현황 영역 조정을 건너뜁니다.", flush=True)
        except Exception as asset_err:
            print(f"[엑셀 병합] 자산목록 파싱 중 에러 발생: {asset_err}", flush=True)
            import traceback
            traceback.print_exc()
 
    # 3. 로드맵 매핑 결과 데이터 기입
    start_row = 26 + diff
    style_source_row = 26 + diff
    M = len(results)
    print(f"[엑셀 병합] 로드맵 매핑 결과 {M}건 기입 및 동적 행 크기 조정 시작", flush=True)
    
    # 로드맵 영역 동적 행 크기 조정
    roadmap_diff = M - 10
    if roadmap_diff > 0:
        sheet.insert_rows(start_row + 10, amount=roadmap_diff)
        for r in range(start_row + 10, start_row + 10 + roadmap_diff):
            for col_idx in range(1, 17):
                src_cell = sheet.cell(row=style_source_row, column=col_idx)
                dest_cell = sheet.cell(row=r, column=col_idx)
                copy_cell_style(src_cell, dest_cell)
    elif roadmap_diff < 0:
        sheet.delete_rows(start_row + M, amount=-roadmap_diff)
    
    for idx, res in enumerate(results):
        current_row = start_row + idx
        for col_idx in range(1, 17):
            src_cell = sheet.cell(row=style_source_row, column=col_idx)
            dest_cell = sheet.cell(row=current_row, column=col_idx)
            copy_cell_style(src_cell, dest_cell)
            
        sheet.cell(row=current_row, column=1, value=idx + 1)
        sheet.cell(row=current_row, column=2, value=res["항목명"])
        sheet.cell(row=current_row, column=3, value=res["세부점검내용"])
        sheet.cell(row=current_row, column=5, value=res["운영현황_증적"])
        sheet.cell(row=current_row, column=7, value=res["개선방안"])
        
        sheet.cell(row=current_row, column=9, value=res["보안영역"])
        sheet.cell(row=current_row, column=10, value=res["과제명"])
        sheet.cell(row=current_row, column=11, value=res["법적요구"])
        sheet.cell(row=current_row, column=12, value=safe_int(res.get("시급성", 3)))
        sheet.cell(row=current_row, column=13, value=safe_int(res.get("위험도", 3)))
        sheet.cell(row=current_row, column=14, value=res["예상예산"])
        sheet.cell(row=current_row, column=15, value=res["로드맵연도"])
 
        # 비고 필드 보정 및 폴백 안전장치 반영
        note_val = res.get("비고", "")
        if not note_val or str(note_val).strip() == "" or str(note_val).strip().lower() == "none":
            rec_sol = res.get("추천솔루션", "N/A")
            rec_vendor = res.get("제조사", "N/A")
            if rec_sol and rec_sol != "N/A":
                note_val = f"[추천 솔루션] {rec_vendor} - {rec_sol}\\n[선정 이유] {rec_sol} 솔루션을 도입하여 고객사의 결함 사항에 대한 조치 및 보안 통제 수준 강화를 권고합니다."
            else:
                note_val = f"[추천 솔루션] N/A\\n[선정 이유] 자사 보안 솔루션 풀 내에 해당하는 기술적 솔루션이 존재하지 않으므로, 장비 도입 대신 사내 보안 지침/절차 수립 및 수동 관리 감독 프로세스 강화를 권고합니다."
        
        sheet.cell(row=current_row, column=16, value=note_val)
 
    # 4. 통합 병합 일괄 적용
    print(f"[엑셀 병합] 통합 병합 복구 시작 (자산목록 오프셋: {diff}, 로드맵 오프셋: {roadmap_diff})", flush=True)
    
    # (1) 백업본 병합 복구
    for m_range in original_merges:
        min_row = m_range.min_row
        max_row = m_range.max_row
        min_col = m_range.min_col
        max_col = m_range.max_col
        
        if min_row < 13:
            sheet.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        elif 13 <= min_row <= 22:
            continue
        elif 23 <= min_row < 26:
            sheet.merge_cells(start_row=min_row + diff, start_column=min_col, end_row=max_row + diff, end_column=max_col)
        elif min_row >= 26:
            if min_row <= 35:
                continue
            sheet.merge_cells(start_row=min_row + diff + roadmap_diff, start_column=min_col, end_row=max_row + diff + roadmap_diff, end_column=max_col)
 
    # (2) 자산목록 세부 현황 신규 병합 적용
    if asset_filepath and os.path.exists(asset_filepath) and N > 0:
        sheet.merge_cells(start_row=12, start_column=14, end_row=12, end_column=16)
        for i in range(N):
            r = 13 + i
            sheet.merge_cells(start_row=r, start_column=14, end_row=r, end_column=16)
            
    # (3) 로드맵 매핑 결과 신규 병합 적용
    for idx in range(M):
        current_row = start_row + idx
        sheet.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
        sheet.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
        sheet.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=8)
        sheet.merge_cells(start_row=current_row, start_column=14, end_row=current_row, end_column=16)
 
    wb.save(output_filepath)
    return output_filepath
"""

# 전체 코드를 합쳐서 새로 작성
full_code = part1_up_to_279 + part2_generate_roadmap_excel

with open("roadmap_agent_llm_new.py", "w", encoding="utf-8") as f:
    f.write(full_code)

print("Restore and update completed successfully!")
