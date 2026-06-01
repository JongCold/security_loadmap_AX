import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
checklist_file = os.path.join(BASE_DIR, "uploads", "20260601_153104_2026년_KG이니시스_내부보안점검_상세체크리스트_v1.4_20260422.xlsx")

wb = openpyxl.load_workbook(checklist_file, data_only=True)
sheet_name = None
for name in wb.sheetnames:
    if "내부보안감사체크리스트" in name:
        sheet_name = name
        break
if not sheet_name:
    sheet_name = wb.sheetnames[3]

sheet = wb[sheet_name]
print(f"Checklist Sheet Name: {sheet.title}")

print("\nMerged cells in checklist (row between 10 and 40):")
merges = sorted(list(sheet.merged_cells.ranges), key=lambda r: r.min_row)
for merge in merges:
    if 10 <= merge.min_row <= 40:
        print(merge)

print("\nDetail of row 11, 24, 29, 35:")
for r in [11, 24, 29, 35]:
    print(f"--- Row {r} ---")
    for c in range(1, 9):
        val = sheet.cell(row=r, column=c).value
        # 병합되었는지 확인하기 위해 merged_cells 탐색
        is_merged = False
        for m in merges:
            if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
                is_merged = True
                merge_range = m
                break
        merge_info = f"(Merged in {merge_range})" if is_merged else ""
        print(f"Col {c} ({openpyxl.utils.get_column_letter(c)}): {val} {merge_info}")
