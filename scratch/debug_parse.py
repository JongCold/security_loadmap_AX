import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

filepath = r"c:\Users\USER\Downloads\Security loadmap_Auto-20260529T052519Z-3-001\Security loadmap_Auto\uploads\20260601_101002_2026년_KG파이낸셜_내부보안점검_상세체크리스트_v0.4_20260424.xlsx"

if not os.path.exists(filepath):
    print("File not found:", filepath)
    sys.exit(1)

wb = openpyxl.load_workbook(filepath, data_only=True)
sheet_name = None
for name in wb.sheetnames:
    if "내부보안감사체크리스트" in name:
        sheet_name = name
        break
if not sheet_name:
    sheet_name = wb.sheetnames[3]

print(f"Using sheet: {sheet_name}")
sheet = wb[sheet_name]

print("\n--- Scan rows with eval_cell is X ---")
count_total_x = 0
for r in range(1, sheet.max_row + 1):
    c3 = sheet.cell(row=r, column=3).value
    c4 = sheet.cell(row=r, column=4).value
    c5 = sheet.cell(row=r, column=5).value
    c6 = sheet.cell(row=r, column=6).value
    c7 = sheet.cell(row=r, column=7).value
    
    c5_str = str(c5).strip().upper() if c5 else ""
    if "X" in c5_str or c5_str == "X":
        count_total_x += 1
        print(f"Row {r} | C3 (항목명): {c3} | C4 (세부점검): {str(c4)[:30]} | C5 (평가): {c5} | C6 (운영현황): {str(c6)[:30]} | C7 (개선): {str(c7)[:30]}")

print(f"\nTotal rows with C5 containing X: {count_total_x}")
