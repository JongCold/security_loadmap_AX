import openpyxl
import os
import sys

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

env_path = os.path.join(BASE_DIR, "KG그룹 정보보안감사_사전환경조사서_양식.xlsx")
asset_path = os.path.join(BASE_DIR, "KG그룹(KG가족사명)자산목록 및 중요도 평가서_양식.xlsx")
template_path = os.path.join(BASE_DIR, "2026년 KG그룹_제로인 정보보안감사_보안솔루션로드맵_v1.2.xlsx")

with open(os.path.join(BASE_DIR, "scratch", "out_utf8.txt"), "w", encoding="utf-8") as out:
    def check_env():
        out.write("\n=== KG그룹 정보보안감사_사전환경조사서_양식.xlsx ===\n")
        wb = openpyxl.load_workbook(env_path, data_only=True)
        out.write(f"Sheets: {wb.sheetnames}\n")
        sheet = wb.worksheets[1] # 2번째 시트
        out.write(f"Sheet 2 Name: {sheet.title}\n")
        for r in range(1, 40):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
            if any(row_vals):
                out.write(f"Row {r}: {row_vals}\n")

    def check_asset():
        out.write("\n=== KG그룹(KG가족사명)자산목록 및 중요도 평가서_양식.xlsx ===\n")
        wb = openpyxl.load_workbook(asset_path, data_only=True)
        out.write(f"Sheets: {wb.sheetnames}\n")
        sheet = wb["4.정보보호시스템"]
        out.write(f"Sheet Name: {sheet.title}\n")
        for r in range(1, 60):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 30)]
            if any(row_vals):
                out.write(f"Row {r}: {row_vals}\n")

    def check_template():
        out.write("\n=== 2026년 KG그룹_제로인 정보보안감사_보안솔루션로드맵_v1.2.xlsx ===\n")
        wb = openpyxl.load_workbook(template_path, data_only=True)
        out.write(f"Sheets: {wb.sheetnames}\n")
        sheet = wb.worksheets[1] # 01_ 시트
        out.write(f"Sheet 2 Name: {sheet.title}\n")
        for r in range(1, 35):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 17)]
            if any(row_vals):
                out.write(f"Row {r}: {row_vals}\n")

    check_env()
    check_asset()
    check_template()
