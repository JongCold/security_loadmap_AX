import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
checklist_file = os.path.join(BASE_DIR, "uploads", "20260601_153104_2026년_KG이니시스_내부보안점검_상세체크리스트_v1.4_20260422.xlsx")

if not os.path.exists(checklist_file):
    # uploads 폴더 내에 다른 이름으로 존재할 수 있으니 찾아봅니다.
    import glob
    files = glob.glob(os.path.join(BASE_DIR, "uploads", "*내부보안점검_상세체크리스트*.xlsx"))
    if files:
        checklist_file = files[0]
    else:
        print("Checklist file not found.")
        exit(1)

print(f"Reading checklist file: {checklist_file}")
wb = openpyxl.load_workbook(checklist_file, data_only=True)
sheet_name = None
for name in wb.sheetnames:
    if "내부보안감사체크리스트" in name:
        sheet_name = name
        break
if not sheet_name:
    sheet_name = wb.sheetnames[3]

sheet = wb[sheet_name]
print(f"Sheet Name: {sheet.title}")

defect_count = 0
last_item_name = ""
for r in range(3, sheet.max_row + 1):
    item_name_cell = sheet.cell(row=r, column=3)
    detail_cell = sheet.cell(row=r, column=4)
    eval_cell = sheet.cell(row=r, column=5)
    status_cell = sheet.cell(row=r, column=6)
    improv_cell = sheet.cell(row=r, column=7)
    
    if item_name_cell.value is not None and str(item_name_cell.value).strip() != "":
        last_item_name = str(item_name_cell.value).strip()
        
    eval_val = str(eval_cell.value).strip().upper() if eval_cell.value else ""
    if eval_val == "X":
        defect_count += 1
        print(f"Defect {defect_count:02d} | Row {r:03d} | Item: {last_item_name} | Detail: {str(detail_cell.value)[:20]} | Eval: {eval_val}")

print(f"Total defects: {defect_count}")
