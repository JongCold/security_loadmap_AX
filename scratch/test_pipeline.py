import os
import sys
import openpyxl

# 프로젝트 루트 경로 추가
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(BASE_DIR)

import roadmap_agent_llm_new as roadmap_agent_llm

print("=== [테스트] 정보보안 로드맵 병합 파이프라인 테스트 시작 ===")

# 테스트 파일 경로 지정
env_file = os.path.join(BASE_DIR, "KG그룹 정보보안감사_사전환경조사서_양식.xlsx")
asset_file = os.path.join(BASE_DIR, "KG그룹(KG가족사명)자산목록 및 중요도 평가서_양식.xlsx")
output_file = os.path.join(BASE_DIR, "scratch", "test_output.xlsx")

# 더미 매핑 결과 데이터
dummy_results = [
    {
        "항목명": "6. 인증 및 권한관리",
        "세부점검내용": "6.1 긴급한 사유 등으로 유지보수 필요 시 외부자에게 부여하는 계정...",
        "운영현황_증적": "임시적 계정발생될 수 있으며...",
        "개선방안": "정보시스템 및 개인정보처리시스템에 대한 접근 시...",
        "보안영역": "Endpoint 보안",
        "과제명": "NAC 도입 필요",
        "법적요구": "필수",
        "시급성": 5,
        "위험도": 5,
        "예상예산": "₩30,000,000",
        "로드맵연도": "2026년",
        "비고": "▶ 추후 NAC에 대한 보안솔루션 계획 수립 필요"
    },
    {
        "항목명": "8. 암호통제",
        "세부점검내용": "8.2 정책에 정의된 암호키의 유효기간에 근거하여 암호키를 주기적으로 변경 관리하고 있는가?",
        "운영현황_증적": "데이터베이스에 들어가 있고 별도 암호키 관리를 하고 있지는 않음",
        "개선방안": "암호키 생성, 이용, 보관, 배포, 파기에 대하여 다음과 같은 내용이 포함된 정책 및 절차를 수립...",
        "보안영역": "시스템 보안",
        "과제명": "DB 암호화 도입 타당성 검토",
        "법적요구": "필수",
        "시급성": 5,
        "위험도": 5,
        "예상예산": "영업 문의가 필요한 영역 논의 필요.",
        "로드맵연도": "2027년",
        "비고": "▶ 추후 DB암호화에 대한 보안솔루션 계획 수립 필요"
    }
]

# 임시 테스트 자산목록 생성 (구분, IP, 벤더 등이 일부 채워진 유효 자산 2건 생성, 1건은 빨간색 배경 적용)
temp_asset_file = os.path.join(BASE_DIR, "scratch", "temp_test_asset.xlsx")
wb_asset = openpyxl.load_workbook(asset_file)
sh_asset = wb_asset["4.정보보호시스템"]

# 9행에 자산 정보 및 빨간색 배경색(운영 미가동 테스트용) 입력, 기밀성3 무결성3 가용성3 설정 (합계9 -> H)
sh_asset.cell(row=9, column=4, value="방화벽")
sh_asset.cell(row=9, column=5, value="192.168.1.10")
sh_asset.cell(row=9, column=6, value="FW-HOST")
sh_asset.cell(row=9, column=7, value="Ahnlab TG-70")
sh_asset.cell(row=9, column=13, value=3)
sh_asset.cell(row=9, column=14, value=3)
sh_asset.cell(row=9, column=15, value=3)
from openpyxl.styles import PatternFill
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
sh_asset.cell(row=9, column=4).fill = red_fill

# 10행에 자산 정보 입력 (배경색 없음), 기밀성2 무결성2 가용성1 설정 (합계5 -> M)
sh_asset.cell(row=10, column=4, value="VPN")
sh_asset.cell(row=10, column=5, value="192.168.1.11")
sh_asset.cell(row=10, column=7, value="FG-60E")
sh_asset.cell(row=10, column=13, value=2)
sh_asset.cell(row=10, column=14, value=2)
sh_asset.cell(row=10, column=15, value=1)

wb_asset.save(temp_asset_file)
print(f"임시 테스트용 자산목록 파일 생성 완료: {temp_asset_file}")

try:
    print(f"1. 사전환경조사서 경로: {env_file} (존재여부: {os.path.exists(env_file)})")
    print(f"2. 임시 자산목록 평가서 경로: {temp_asset_file} (존재여부: {os.path.exists(temp_asset_file)})")
    print("3. 엑셀 추출 기동...")
    
    roadmap_agent_llm.generate_roadmap_excel(
        results=dummy_results,
        company_name="테스트고객사",
        output_filepath=output_file,
        env_filepath=env_file,
        asset_filepath=temp_asset_file
    )
    
    print(f"✅ 테스트 성공! 최종 로드맵 파일 생성 완료: {output_file}")
    
    # 생성된 파일의 시트 구조 검사
    wb = openpyxl.load_workbook(output_file, data_only=False) # 스타일 속성을 보기 위해 False
    sheet = wb.worksheets[1]
    print(f"Sheet title: {sheet.title}")
    print(f"B5 (기업명): {sheet.cell(row=5, column=2).value}")
    print(f"I5 (사업분야): {sheet.cell(row=5, column=9).value}")
    print("Row 12 Headers (9~14열):", [sheet.cell(row=12, column=col).value for col in range(9, 15)])
    print("Row 12 Column 9 (기밀성 헤더) 정렬 방식:", sheet.cell(row=12, column=9).alignment.horizontal)
    print("Row 12 Column 9 (기밀성 헤더) 배경색 (Style Copy 검증):", sheet.cell(row=12, column=9).fill.start_color.rgb)
    print(f"Row 13 (첫 번째 자산 데이터 1~14열): {[sheet.cell(row=13, column=c).value for c in range(1, 15)]}")
    print(f"Row 13 구분 셀 배경색 (Style Copy 검증): {sheet.cell(row=13, column=2).fill.start_color.rgb}")
    print(f"Row 13 Column 9 (기밀성 데이터) 정렬 방식:", sheet.cell(row=13, column=9).alignment.horizontal)
    print(f"Row 13 Column 9 (기밀성 데이터) 테두리 스타일:", sheet.cell(row=13, column=9).border.bottom.style)
    print(f"Row 14 (두 번째 자산 데이터 1~14열): {[sheet.cell(row=14, column=c).value for c in range(1, 15)]}")
    print(f"Row 14 구분 셀 배경색 (Style Copy 검증 - 배경 무): {sheet.cell(row=14, column=2).fill.fill_type}")
    print(f"Row 13 높이: {sheet.row_dimensions[13].height}")
    print(f"Row 14 높이: {sheet.row_dimensions[14].height}")
    
    # N = 2 이므로 diff = -8, 로드맵 시작 위치는 26 - 8 = 18행이어야 함
    print(f"Row 18 (첫 번째 로드맵 데이터): {[sheet.cell(row=18, column=c).value for c in range(1, 17)]}")
    print(f"Row 19 (두 번째 로드맵 데이터): {[sheet.cell(row=19, column=c).value for c in range(1, 17)]}")
    
except Exception as e:
    print(f"❌ 에러 발생: {e}")
    import traceback
    traceback.print_exc()
    
# 임시 테스트 파일 정리
if os.path.exists(temp_asset_file):
    os.remove(temp_asset_file)
